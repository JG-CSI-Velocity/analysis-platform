"""Generate slide_catalog.xlsx -- one tab per pipeline with every analysis/slide."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

COLUMNS = [
    "Section",
    "Module",
    "Slide / Analysis ID",
    "Title",
    "Chart Type",
    "Layout Index",
    "Slide Type",
    "Data Source",
    "Notes",
]

ARS_DATA = [
    # Overview
    ["Overview", "stat_codes", "A1", "Stat Code Distribution", "Bar", 9, "screenshot", "ODD", "Full-width bar; many categories"],
    ["Overview", "product_codes", "A1b", "Product Code Distribution", "Bar", 5, "chart_narrative", "ODD", "Chart + product mix commentary"],
    ["Overview", "eligibility", "A3", "Eligibility Analysis", "Scatter", 5, "chart_narrative", "ODD", "Scatter + eligibility criteria explanation"],
    # DCTR - Penetration
    ["DCTR", "penetration", "DCTR-1", "Historical Debit Card Take Rate", "Line", 5, "chart_narrative", "ODD", "Trend line + rate narrative; stored in ctx.results"],
    ["DCTR", "penetration", "DCTR-2", "DCTR: Open vs Eligible", "Grouped Bar", 6, "comparison", "ODD", "Side-by-side comparison of two populations"],
    ["DCTR", "penetration", "DCTR-3", "DCTR Snapshot: Open to TTM", "Grouped Bar", 4, "chart_kpi", "ODD", "Chart + KPI callouts for current vs TTM"],
    ["DCTR", "penetration", "DCTR-4", "Personal DCTR", "Line", 6, "comparison", "ODD", "Pair with DCTR-5 (Personal vs Business side-by-side)"],
    ["DCTR", "penetration", "DCTR-5", "Business DCTR", "Line", 6, "comparison", "ODD", "Pair with DCTR-4 (Personal vs Business side-by-side)"],
    ["DCTR", "penetration", "DCTR-6", "Personal L12M DCTR", "Line", 6, "comparison", "ODD", "Pair with DCTR-7 (L12M Personal vs Business)"],
    ["DCTR", "penetration", "DCTR-7", "Business L12M DCTR", "Line", 6, "comparison", "ODD", "Pair with DCTR-6 (L12M Personal vs Business)"],
    ["DCTR", "penetration", "DCTR-8", "Comprehensive DCTR Summary", "Grouped Bar", 9, "screenshot", "ODD", "Full-width; dense multi-group data"],
    # DCTR - Trends
    ["DCTR", "trends", "A7.4", "DCTR Recent Trend", "Line", 5, "chart_narrative", "ODD", "Trend line + narrative on trajectory; reads penetration results"],
    ["DCTR", "trends", "A7.5", "DCTR Segment Analysis", "Grouped Bar", 9, "screenshot", "ODD", "Full-width; many segments x groups"],
    ["DCTR", "trends", "A7.6a", "DCTR Trajectory + Segments", "Multi-chart", 6, "multi_screenshot", "ODD", "Merged pair (trajectory left, segments right)"],
    ["DCTR", "trends", "A7.6b", "DCTR Segment Breakdown", "Bar", 5, "chart_narrative", "ODD", "Chart + segment insight callout"],
    ["DCTR", "trends", "A7.14", "DCTR Cohort Analysis", "Line", 5, "chart_narrative", "ODD", "Cohort curves + cohort explanation text"],
    ["DCTR", "trends", "A7.15", "DCTR Cohort Detail", "Grouped Bar", 9, "screenshot", "ODD", "Full-width detail breakdown"],
    # DCTR - Branches
    ["DCTR", "branches", "DCTR-9", "Branch DCTR Comparison", "Heatmap", 9, "screenshot", "ODD", "Full-width heatmap; many branches x periods"],
    ["DCTR", "branches", "A7.10a", "Branch Performance Matrix", "Heatmap", 13, "wide_custom", "ODD", "Wide layout; dense branch matrix"],
    ["DCTR", "branches", "A7.10b", "Branch Detail View", "Bar", 5, "chart_narrative", "ODD", "Bar + branch performance commentary"],
    ["DCTR", "branches", "A7.10c", "Branch KPI Summary", "KPI Panel", 7, "kpi_grid", "ODD", "6-cell KPI grid for branch metrics"],
    ["DCTR", "branches", "DCTR-15", "Branch Rank", "Bar", 6, "comparison", "ODD", "Pair with DCTR-16 (rank left, change right)"],
    ["DCTR", "branches", "DCTR-16", "Branch Change", "Lollipop", 6, "comparison", "ODD", "Pair with DCTR-15 (rank left, change right)"],
    ["DCTR", "branches", "A7.13", "Branch Appendix Detail", "Heatmap", 9, "screenshot", "ODD", "Appendix; full-width reference heatmap"],
    # DCTR - Funnel
    ["DCTR", "funnel", "A7.7", "DCTR Funnel: Historical vs TTM", "Multi-chart", 6, "multi_screenshot", "ODD", "Merged pair (historical left, TTM right)"],
    ["DCTR", "funnel", "A7.8", "Account Eligibility Funnel", "Funnel", 5, "chart_narrative", "ODD", "Funnel chart + stage-by-stage explanation"],
    ["DCTR", "funnel", "A7.9", "Funnel Detail Breakdown", "Bar", 9, "screenshot", "ODD", "Full-width detail bar"],
    # DCTR - Overlays
    ["DCTR", "overlays", "DCTR-10", "Age Distribution Overlay", "Histogram", 5, "chart_narrative", "ODD", "Histogram + age distribution insights"],
    ["DCTR", "overlays", "DCTR-11", "Product Code Breakdown", "Donut", 5, "chart_narrative", "ODD", "Donut + product mix narrative"],
    ["DCTR", "overlays", "DCTR-12", "Personal vs Business", "Grouped Bar", 6, "comparison", "ODD", "Natural comparison; split layout"],
    ["DCTR", "overlays", "DCTR-13", "Balance Tier Analysis", "Grouped Bar", 5, "chart_narrative", "ODD", "Chart + tier threshold explanation"],
    ["DCTR", "overlays", "DCTR-14", "Account Age Impact", "Scatter", 5, "chart_narrative", "ODD", "Scatter + correlation narrative"],
    # Reg E - Status
    ["Reg E", "status", "A8.1", "Reg E Overall Rate", "Line", 5, "chart_narrative", "ODD", "Trend line + rate trajectory narrative"],
    ["Reg E", "status", "A8.2", "Reg E: Open vs Eligible", "Grouped Bar", 6, "comparison", "ODD", "Side-by-side comparison"],
    ["Reg E", "status", "A8.3", "Reg E Snapshot", "Grouped Bar", 4, "chart_kpi", "ODD", "Chart + 2 KPI callout zones"],
    ["Reg E", "status", "A8.12", "Reg E Summary KPI", "KPI", 7, "kpi_grid", "ODD", "6-cell KPI grid for Reg E metrics"],
    # Reg E - Dimensions
    ["Reg E", "dimensions", "A8.5", "Reg E Opportunity: Age", "Multi-chart", 6, "multi_screenshot", "ODD", "Merged pair (opportunity left, detail right)"],
    ["Reg E", "dimensions", "A8.6", "Reg E by Age Detail", "Bar", 5, "chart_narrative", "ODD", "Chart + age-band commentary"],
    ["Reg E", "dimensions", "A8.7", "Reg E by Tenure", "Heatmap", 9, "screenshot", "ODD", "Full-width heatmap; many tenure bands"],
    ["Reg E", "dimensions", "A8.10", "Reg E Funnel", "Multi-chart", 6, "multi_screenshot", "ODD", "Merged pair"],
    ["Reg E", "dimensions", "A8.11", "Reg E Funnel Detail", "Funnel", 5, "chart_narrative", "ODD", "Funnel + stage explanation"],
    # Reg E - Branches
    ["Reg E", "branches", "A8.4a", "Branch Reg E Matrix", "Heatmap", 13, "wide_custom", "ODD", "Wide layout; dense branch matrix"],
    ["Reg E", "branches", "A8.4b", "Branch Reg E Detail", "Bar", 5, "chart_narrative", "ODD", "Bar + branch commentary"],
    ["Reg E", "branches", "A8.4c", "Branch Appendix", "Heatmap", 9, "screenshot", "ODD", "Appendix; full-width reference"],
    ["Reg E", "branches", "A8.13", "Reg E Branch Rank", "Bar", 9, "screenshot", "ODD", "Full-width rank bar; many branches"],
    # Attrition - Rates
    ["Attrition", "rates", "A9.1", "Overall Attrition Rate", "KPI + Bar", 4, "chart_kpi", "ODD", "Chart + 2 KPI zones (rate + YoY change)"],
    ["Attrition", "rates", "A9.2", "Closure Duration", "Horizontal Bar", 5, "chart_narrative", "ODD", "Hbar + duration distribution insights"],
    ["Attrition", "rates", "A9.3", "Open vs Closed", "Multi-chart", 6, "multi_screenshot", "ODD", "Merged pair (open left, closed right)"],
    # Attrition - Dimensions
    ["Attrition", "dimensions", "A9.4", "Closure by Branch", "Horizontal Bar", 9, "screenshot", "ODD", "Full-width; many branches"],
    ["Attrition", "dimensions", "A9.5", "Closure by Product", "Bar", 5, "chart_narrative", "ODD", "Chart + product attrition narrative"],
    ["Attrition", "dimensions", "A9.6", "Personal vs Business", "Multi-chart", 6, "multi_screenshot", "ODD", "Merged pair (personal left, business right)"],
    ["Attrition", "dimensions", "A9.7", "Closure by Tenure", "Bar", 5, "chart_narrative", "ODD", "Chart + tenure band insights"],
    ["Attrition", "dimensions", "A9.8", "Closure by Balance", "Bar", 5, "chart_narrative", "ODD", "Chart + balance tier insights"],
    # Attrition - Impact
    ["Attrition", "impact", "A9.9", "Debit Retention Impact", "KPI + Chart", 4, "chart_kpi", "ODD", "Chart + retention rate + revenue KPIs"],
    ["Attrition", "impact", "A9.10", "Mailer Retention Impact", "KPI + Chart", 4, "chart_kpi", "ODD", "Chart + mailer lift + response KPIs"],
    ["Attrition", "impact", "A9.11", "Revenue Lost", "KPI + Chart", 4, "chart_kpi", "ODD", "Chart + lost revenue + account count KPIs"],
    ["Attrition", "impact", "A9.12", "L12M Velocity Impact", "KPI + Line", 5, "chart_narrative", "ODD", "Trend line + velocity narrative"],
    ["Attrition", "impact", "A9.13", "ARS vs Non-ARS Attrition", "Bar", 6, "comparison", "ODD", "Natural comparison; ARS left vs Non-ARS right"],
    # Value
    ["Value", "analysis", "A11.1", "Value of Debit Card", "Waterfall", 5, "chart_narrative", "ODD", "Waterfall + value buildup narrative"],
    ["Value", "analysis", "A11.2", "Value of Reg E", "Waterfall", 5, "chart_narrative", "ODD", "Waterfall + Reg E value narrative"],
    # Mailer - Insights (monthly, dynamic count)
    ["Mailer", "insights", "A13.{month}", "Monthly Response Summary", "Composite", 7, "kpi_grid", "ODD", "6-cell KPI grid; one per mailer month"],
    ["Mailer", "insights", "A12.{month}.Swipes", "Monthly Swipes Trend", "Line", 6, "comparison", "ODD", "Pair with Spend (swipes left, spend right)"],
    ["Mailer", "insights", "A12.{month}.Spend", "Monthly Spend Trend", "Line", 6, "comparison", "ODD", "Pair with Swipes (swipes left, spend right)"],
    # Mailer - Response
    ["Mailer", "response", "A13.5", "Program Response Count Trend", "Line", 5, "chart_narrative", "ODD", "Trend line + response trajectory text"],
    ["Mailer", "response", "A13.6", "Response Rate Trend", "Line", 5, "chart_narrative", "ODD", "Trend line + rate commentary"],
    ["Mailer", "response", "A13.Agg", "Aggregate Program Summary", "Donut/Bar", 7, "kpi_grid", "ODD", "Multi-KPI aggregate summary"],
    ["Mailer", "response", "A14.2", "Mailer Revisit Analysis", "Bar", 5, "chart_narrative", "ODD", "Chart + revisit pattern narrative"],
    # Mailer - Impact
    ["Mailer", "impact", "A15.1", "Mailer Lift Revenue", "KPI", 4, "chart_kpi", "ODD", "Chart + revenue lift + ROI KPIs"],
    ["Mailer", "impact", "A15.2", "Mailer Response Driver", "Bar", 5, "chart_narrative", "ODD", "Chart + driver analysis text"],
    ["Mailer", "impact", "A15.3", "Mailer Effectiveness", "Line", 5, "chart_narrative", "ODD", "Trend line + effectiveness narrative"],
    ["Mailer", "impact", "A15.4", "Mailer ROI Summary", "KPI", 7, "kpi_grid", "ODD", "6-cell KPI grid for ROI metrics"],
    # Insights
    ["Insights", "synthesis", "S1", "Growth Drivers", "Text + Chart", 5, "chart_narrative", "ODD", "Chart + growth driver narrative"],
    ["Insights", "synthesis", "S2", "Risk Factors", "Text + Chart", 5, "chart_narrative", "ODD", "Chart + risk factor narrative"],
    ["Insights", "synthesis", "S3", "Opportunity Analysis", "Text + Chart", 5, "chart_narrative", "ODD", "Chart + opportunity narrative"],
    ["Insights", "synthesis", "S4", "Program Impact", "Text + Chart", 5, "chart_narrative", "ODD", "Chart + program impact narrative"],
    ["Insights", "synthesis", "S5", "Recommendations", "Text + Chart", 5, "chart_narrative", "ODD", "Chart + actionable recommendations"],
    ["Insights", "conclusions", "S6", "Conclusion 1", "Text", 2, "section_text", "ODD", "Divider layout with conclusion text in content area"],
    ["Insights", "conclusions", "S7", "Conclusion 2", "Text", 2, "section_text", "ODD", "Divider layout with conclusion text in content area"],
    ["Insights", "conclusions", "S8", "Conclusion 3", "Text", 2, "section_text", "ODD", "Divider layout with conclusion text in content area"],
]

TXN_DATA = [
    # M1: Overall
    ["M1 Overall", "overall", "top_merchants_by_spend", "Top Merchants by Spend", "Lollipop", 9, "screenshot", "TXN CSV", "Full-width; top 25 merchants"],
    ["M1 Overall", "overall", "top_merchants_by_transactions", "Top Merchants by Transactions", "Lollipop", 9, "screenshot", "TXN CSV", "Full-width; top 25 merchants"],
    ["M1 Overall", "overall", "top_merchants_by_accounts", "Top Merchants by Accounts", "Lollipop", 9, "screenshot", "TXN CSV", "Full-width; top 25 merchants"],
    # M2: MCC
    ["M2 MCC", "mcc", "mcc_by_accounts", "MCC by Accounts", "Table", "", "data_only", "TXN CSV", "Excel-only data table; no chart"],
    ["M2 MCC", "mcc", "mcc_by_transactions", "MCC by Transactions", "Table", "", "data_only", "TXN CSV", "Excel-only data table; no chart"],
    ["M2 MCC", "mcc", "mcc_by_spend", "MCC by Spend", "Table", "", "data_only", "TXN CSV", "Excel-only data table; no chart"],
    ["M2 MCC", "mcc", "mcc_comparison", "MCC Comparison", "3-Panel Bar", 8, "grid_thirds", "TXN CSV", "3 panels in 3-col grid; needs all 3 MCC results"],
    # M3: Business
    ["M3 Business", "business", "business_top_by_spend", "Business Top by Spend", "Lollipop", 9, "screenshot", "TXN CSV", "Full-width; many merchants"],
    ["M3 Business", "business", "business_top_by_transactions", "Business Top by Transactions", "Lollipop", 9, "screenshot", "TXN CSV", "Full-width; many merchants"],
    ["M3 Business", "business", "business_top_by_accounts", "Business Top by Accounts", "Lollipop", 9, "screenshot", "TXN CSV", "Full-width; many merchants"],
    # M4: Personal
    ["M4 Personal", "personal", "personal_top_by_spend", "Personal Top by Spend", "Lollipop", 9, "screenshot", "TXN CSV", "Full-width; #62 decompression bomb fixed"],
    ["M4 Personal", "personal", "personal_top_by_transactions", "Personal Top by Transactions", "Lollipop", 9, "screenshot", "TXN CSV", "Full-width; many merchants"],
    ["M4 Personal", "personal", "personal_top_by_accounts", "Personal Top by Accounts", "Lollipop", 9, "screenshot", "TXN CSV", "Full-width; many merchants"],
    # M5: Trends
    ["M5 Trends", "trends", "monthly_rank_tracking", "Monthly Rank Trajectory", "Line", 5, "chart_narrative", "TXN CSV", "Trajectory line + rank movement commentary"],
    ["M5 Trends", "trends", "growth_leaders_decliners", "Growth Leaders & Decliners", "Grouped Bar", 4, "chart_kpi", "TXN CSV", "Chart + top leader/decliner KPI callouts"],
    ["M5 Trends", "trends", "spending_consistency", "Spending Consistency", "Scatter", 5, "chart_narrative", "TXN CSV", "Scatter + consistency explanation; no chart yet"],
    ["M5 Trends", "trends", "new_vs_declining_merchants", "New vs Declining Merchants", "Bar", 6, "comparison", "TXN CSV", "Natural split: new left vs declining right"],
    ["M5 Trends", "trends", "business_monthly_movers", "Business Monthly Movers", "Table", "", "data_only", "TXN CSV", "Excel-only table"],
    ["M5 Trends", "trends", "personal_monthly_movers", "Personal Monthly Movers", "Table", "", "data_only", "TXN CSV", "Excel-only table"],
    # M6: Competitor
    ["M6 Competitor", "competitor", "competitor_detection", "Competitor Detection", "Data", "", "data_only", "TXN CSV", "Engine step; populates context for downstream"],
    ["M6 Competitor", "competitor", "competitor_high_level", "Competitor High Level", "KPI", 7, "kpi_grid", "TXN CSV", "6-cell KPI grid: share, count, trend metrics"],
    ["M6 Competitor", "competitor", "top_20_competitors", "Top 20 Competitors", "Table", "", "data_only", "TXN CSV", "Excel-only table"],
    ["M6 Competitor", "competitor", "competitor_categories", "Competitor Categories", "Heatmap", 9, "screenshot", "TXN CSV", "Full-width heatmap; many categories x competitors"],
    ["M6 Competitor", "competitor", "competitor_biz_personal", "Competitor Biz vs Personal", "Table", 6, "comparison", "TXN CSV", "Split: business left vs personal right"],
    ["M6 Competitor", "competitor", "competitor_monthly_trends", "Competitor Monthly Trends", "Line", 5, "chart_narrative", "TXN CSV", "Trend line + competitive shift narrative"],
    ["M6 Competitor", "competitor", "competitor_threat_assessment", "Competitor Threat Assessment", "Scatter", 5, "chart_narrative", "TXN CSV", "Scatter + threat level explanation"],
    ["M6 Competitor", "competitor", "competitor_segmentation", "Competitor Segmentation", "Bar", 5, "chart_narrative", "TXN CSV", "Chart + segmentation breakdown text"],
    ["M6 Competitor", "competitor", "unmatched_financial", "Unmatched Financial", "Table", "", "data_only", "TXN CSV", "Data quality audit; Excel-only"],
    # M7: Financial
    ["M7 Financial", "financial", "financial_services_detection", "Financial Services Detection", "Data", "", "data_only", "TXN CSV", "Engine step; must run before M7B"],
    ["M7 Financial", "financial", "financial_services_summary", "Financial Services Summary", "KPI", 7, "kpi_grid", "TXN CSV", "6-cell KPI grid for financial services metrics"],
    # M8: Interchange
    ["M8 Interchange", "interchange", "interchange_summary", "Interchange Summary", "KPI", 7, "kpi_grid", "TXN CSV", "6-cell KPI grid for interchange metrics"],
    # M10: Member
    ["M10 Member", "member", "member_segments", "Member Segments", "KPI", 7, "kpi_grid", "TXN CSV", "6-cell KPI grid for segment breakdown"],
    # M11: Demographics
    ["M11 Demographics", "demographics", "demographics", "Demographics (Generation)", "Table/Charts", 5, "chart_narrative", "TXN + ODD", "Chart + demographic narrative; requires ODD generation col"],
    # M15: Recurring
    ["M15 Recurring", "recurring", "recurring_payments", "Recurring Payment Merchants", "Bar", 5, "chart_narrative", "TXN CSV", "Chart + recurring merchant narrative"],
    ["M15 Recurring", "recurring", "recurring_payments:onsets", "Recurring Merchant Onsets", "Line", 5, "chart_narrative", "TXN CSV", "Trend line + onset timing narrative"],
    # M9: Scorecard
    ["M9 Scorecard", "scorecard", "portfolio_scorecard", "Portfolio Scorecard", "Bullet Chart", 8, "grid_thirds", "TXN CSV", "6-cell grid of bullet gauges; must run last"],
]

ICS_DATA = [
    # Summary
    ["Summary", "summary", "ax01", "Total ICS Accounts", "KPI Gauge", 7, "kpi_grid", "ICS Excel", "6-cell KPI grid: total, open, closed, rate"],
    ["Summary", "summary", "ax02", "Open ICS Accounts", "KPI", 7, "kpi_grid", "ICS Excel", "Combine with ax01 on same KPI grid"],
    ["Summary", "summary", "ax07", "ICS by Stat Code", "Bar", 9, "screenshot", "ICS Excel", "Full-width; many stat codes"],
    ["Summary", "summary", "ax06", "Product Code Distribution", "Donut", 5, "chart_narrative", "ICS Excel", "Donut + product mix commentary"],
    ["Summary", "summary", "ax05", "Debit Distribution", "Donut", 5, "chart_narrative", "ICS Excel", "Donut + debit penetration narrative"],
    ["Summary", "summary", "ax64", "Debit x Product Code", "Grouped Bar", 9, "screenshot", "ICS Excel", "Full-width; multi-group cross-tab"],
    ["Summary", "summary", "ax04", "Debit x Branch", "Heatmap", 9, "screenshot", "ICS Excel", "Full-width heatmap; many branches"],
    ["Summary", "summary", "ax03", "ICS Penetration by Branch", "Bar", 5, "chart_narrative", "ICS Excel", "Chart + penetration rate narrative"],
    # Source
    ["Source", "source", "ax08", "Source Distribution", "Donut", 5, "chart_narrative", "ICS Excel", "Donut + REF/DM/Both breakdown narrative"],
    ["Source", "source", "ax85", "Source x Stat Code", "Stacked Bar", 9, "screenshot", "ICS Excel", "Full-width; complex stacked breakdown"],
    ["Source", "source", "ax09", "Source x Product Code", "Grouped Bar", 9, "screenshot", "ICS Excel", "Full-width; multi-group cross-tab"],
    ["Source", "source", "ax10", "Source x Branch", "Heatmap", 9, "screenshot", "ICS Excel", "Full-width heatmap; many branches x sources"],
    ["Source", "source", "ax11", "Account Type", "Donut", 6, "comparison", "ICS Excel", "Split: business left vs personal right"],
    ["Source", "source", "ax12", "Source by Year", "Grouped Bar", 5, "chart_narrative", "ICS Excel", "Chart + year-over-year acquisition narrative"],
    ["Source", "source", "ax13", "Source Acquisition Mix", "Waterfall", 5, "chart_narrative", "ICS Excel", "Waterfall + acquisition channel explanation"],
    # DM Deep-Dive
    ["DM Deep-Dive", "dm_source", "ax45", "DM Overview", "Table", "", "data_only", "ICS Excel", "Excel-only summary table"],
    ["DM Deep-Dive", "dm_source", "ax46", "DM by Branch", "Heatmap", 9, "screenshot", "ICS Excel", "Full-width heatmap; many branches"],
    ["DM Deep-Dive", "dm_source", "ax47", "DM by Debit Status", "Grouped Bar", 5, "chart_narrative", "ICS Excel", "Chart + DM debit activation narrative"],
    ["DM Deep-Dive", "dm_source", "ax48", "DM by Product", "Grouped Bar", 5, "chart_narrative", "ICS Excel", "Chart + DM product mix narrative"],
    ["DM Deep-Dive", "dm_source", "ax49", "DM by Year", "Line", 5, "chart_narrative", "ICS Excel", "Trend line + DM acquisition trajectory"],
    ["DM Deep-Dive", "dm_source", "ax50", "DM Activity Summary", "KPI", 7, "kpi_grid", "ICS Excel", "6-cell KPI grid for DM activity metrics"],
    ["DM Deep-Dive", "dm_source", "ax51", "DM Activity by Branch", "Grouped Bar", 9, "screenshot", "ICS Excel", "Full-width; many branches"],
    ["DM Deep-Dive", "dm_source", "ax52", "DM Monthly Trends", "Line", 5, "chart_narrative", "ICS Excel", "Trend line + monthly trajectory narrative"],
    # REF Deep-Dive
    ["REF Deep-Dive", "ref_source", "ax73", "REF Overview", "Table", "", "data_only", "ICS Excel", "Excel-only summary table"],
    ["REF Deep-Dive", "ref_source", "ax74", "REF by Branch", "Heatmap", 9, "screenshot", "ICS Excel", "Full-width heatmap; many branches"],
    ["REF Deep-Dive", "ref_source", "ax75", "REF by Debit Status", "Grouped Bar", 5, "chart_narrative", "ICS Excel", "Chart + REF debit activation narrative"],
    ["REF Deep-Dive", "ref_source", "ax76", "REF by Product", "Grouped Bar", 5, "chart_narrative", "ICS Excel", "Chart + REF product mix narrative"],
    ["REF Deep-Dive", "ref_source", "ax77", "REF by Year", "Line", 5, "chart_narrative", "ICS Excel", "Trend line + REF acquisition trajectory"],
    ["REF Deep-Dive", "ref_source", "ax78", "REF Activity Summary", "KPI", 7, "kpi_grid", "ICS Excel", "6-cell KPI grid for REF activity metrics"],
    ["REF Deep-Dive", "ref_source", "ax79", "REF Activity by Branch", "Grouped Bar", 9, "screenshot", "ICS Excel", "Full-width; many branches"],
    ["REF Deep-Dive", "ref_source", "ax80", "REF Monthly Trends", "Line", 5, "chart_narrative", "ICS Excel", "Trend line + monthly trajectory narrative"],
    # Demographics
    ["Demographics", "demographics", "ax14", "Age Comparison", "Grouped Bar", 6, "comparison", "ICS Excel", "Split: ICS vs non-ICS age distributions"],
    ["Demographics", "demographics", "ax15", "Closures", "Donut", 5, "chart_narrative", "ICS Excel", "Donut + closure pattern narrative"],
    ["Demographics", "demographics", "ax16", "Open vs Close", "Grouped Bar", 6, "comparison", "ICS Excel", "Natural split: open left vs closed right"],
    ["Demographics", "demographics", "ax17", "Balance Tiers", "Grouped Bar", 5, "chart_narrative", "ICS Excel", "Chart + balance tier distribution narrative"],
    ["Demographics", "demographics", "ax83", "Stat Open Close", "Heatmap", 9, "screenshot", "ICS Excel", "Full-width heatmap; stat codes x open/close"],
    ["Demographics", "demographics", "ax18", "Age vs Balance", "Scatter", 5, "chart_narrative", "ICS Excel", "Scatter + age-balance correlation narrative"],
    ["Demographics", "demographics", "ax19", "Balance Tier Detail", "Grouped Bar", 9, "screenshot", "ICS Excel", "Full-width; many tiers x breakdowns"],
    ["Demographics", "demographics", "ax20", "Age Distribution", "Histogram", 5, "chart_narrative", "ICS Excel", "Histogram + age distribution insights"],
    ["Demographics", "demographics", "ax21", "Balance Trajectory", "Line", 5, "chart_narrative", "ICS Excel", "Trend line + balance trajectory narrative"],
    # Activity
    ["Activity", "activity", "ax22", "Activity Summary", "KPI", 7, "kpi_grid", "ICS Excel", "6-cell KPI grid: active, dormant, rates"],
    ["Activity", "activity", "ax23", "Activity by Debit+Source", "Grouped Bar", 9, "screenshot", "ICS Excel", "Full-width; multi-factor cross-tab"],
    ["Activity", "activity", "ax24", "Activity by Balance", "Grouped Bar", 5, "chart_narrative", "ICS Excel", "Chart + balance-activity correlation narrative"],
    ["Activity", "activity", "ax25", "Activity by Branch", "Grouped Bar", 9, "screenshot", "ICS Excel", "Full-width; many branches"],
    ["Activity", "activity", "ax63", "Monthly Trends", "Line", 5, "chart_narrative", "ICS Excel", "Trend line + monthly activity narrative"],
    ["Activity", "activity", "ax71", "Activity Source Comparison", "Line", 6, "comparison", "ICS Excel", "Split: DM trend left vs REF trend right"],
    ["Activity", "activity", "ax72", "Monthly Interchange", "Bar", 5, "chart_narrative", "ICS Excel", "Chart + interchange revenue narrative"],
    ["Activity", "activity", "ax26", "Business vs Personal", "Grouped Bar", 6, "comparison", "ICS Excel", "Split: business left vs personal right"],
    # Cohort
    ["Cohort", "cohort", "ax27", "Cohort Activation", "Heatmap", 9, "screenshot", "ICS Excel", "Full-width heatmap; cohorts x months"],
    ["Cohort", "cohort", "ax28", "Cohort Heatmap", "Heatmap", 9, "screenshot", "ICS Excel", "Full-width heatmap; cohorts x metrics"],
    ["Cohort", "cohort", "ax29", "Cohort Milestones", "Line", 5, "chart_narrative", "ICS Excel", "Milestone curves + activation timeline narrative"],
    ["Cohort", "cohort", "ax30", "Activation Summary", "KPI", 7, "kpi_grid", "ICS Excel", "6-cell KPI grid: activation rates by cohort"],
    ["Cohort", "cohort", "ax31", "Growth Patterns", "Line", 5, "chart_narrative", "ICS Excel", "Growth curves + pattern narrative"],
    ["Cohort", "cohort", "ax32", "Activation Personas", "Grouped Bar", 5, "chart_narrative", "ICS Excel", "Chart + persona profile narrative"],
    ["Cohort", "cohort", "ax33", "Branch Activation", "Bar", 9, "screenshot", "ICS Excel", "Full-width; many branches"],
    # Strategic
    ["Strategic", "strategic", "ax38", "Activation Funnel", "Funnel", 5, "chart_narrative", "ICS Excel", "Funnel + stage-by-stage explanation"],
    ["Strategic", "strategic", "ax39", "Revenue Impact", "Bar", 4, "chart_kpi", "ICS Excel", "Chart + revenue KPI callouts"],
    ["Strategic", "strategic", "ax65", "Revenue by Branch", "Bar", 9, "screenshot", "ICS Excel", "Full-width; many branches"],
    ["Strategic", "strategic", "ax66", "Revenue by Source", "Donut", 5, "chart_narrative", "ICS Excel", "Donut + source revenue narrative"],
    ["Strategic", "strategic", "ax84", "Dormant High-Balance", "Table", "", "data_only", "ICS Excel", "Action list; Excel-only table"],
    # Portfolio
    ["Portfolio", "portfolio", "ax40", "Engagement Decay", "Line", 5, "chart_narrative", "ICS Excel", "Decay curve + retention strategy narrative"],
    ["Portfolio", "portfolio", "ax41", "Net Portfolio Growth", "Line", 5, "chart_narrative", "ICS Excel", "Growth line + portfolio health narrative"],
    ["Portfolio", "portfolio", "ax42", "Spend Concentration", "Scatter", 5, "chart_narrative", "ICS Excel", "Scatter + concentration risk narrative"],
    ["Portfolio", "portfolio", "ax67", "Closure by Source", "Bar", 5, "chart_narrative", "ICS Excel", "Chart + source attrition narrative"],
    ["Portfolio", "portfolio", "ax68", "Closure by Branch", "Bar", 9, "screenshot", "ICS Excel", "Full-width; many branches"],
    ["Portfolio", "portfolio", "ax69", "Closure by Account Age", "Line", 5, "chart_narrative", "ICS Excel", "Line + age-attrition correlation narrative"],
    ["Portfolio", "portfolio", "ax70", "Net Growth by Source", "Bar", 5, "chart_narrative", "ICS Excel", "Chart + source growth narrative"],
    ["Portfolio", "portfolio", "ax82", "Closure Rate Trend", "Line", 5, "chart_narrative", "ICS Excel", "Trend line + closure rate narrative"],
    # Performance
    ["Performance", "performance", "ax43", "Days to First Use", "Histogram", 5, "chart_narrative", "ICS Excel", "Histogram + activation speed narrative"],
    ["Performance", "performance", "ax44", "Branch Performance Index", "Bar", 9, "screenshot", "ICS Excel", "Full-width; ranked branches"],
    ["Performance", "performance", "ax81", "Product Code Performance", "Bar", 9, "screenshot", "ICS Excel", "Full-width; ranked products"],
    # Persona
    ["Persona", "persona", "ax55", "Persona Overview", "Bubble", 5, "chart_narrative", "ICS Excel", "Bubble chart + persona definition narrative"],
    ["Persona", "persona", "ax56", "Persona Contribution", "Bar", 5, "chart_narrative", "ICS Excel", "Chart + contribution breakdown narrative"],
    ["Persona", "persona", "ax57", "Persona by Branch", "Heatmap", 9, "screenshot", "ICS Excel", "Full-width heatmap; personas x branches"],
    ["Persona", "persona", "ax58", "Persona by Source", "Grouped Bar", 5, "chart_narrative", "ICS Excel", "Chart + source-persona narrative"],
    ["Persona", "persona", "ax59", "Persona Revenue", "Bar", 5, "chart_narrative", "ICS Excel", "Chart + per-persona revenue narrative"],
    ["Persona", "persona", "ax60", "Persona by Balance", "Table", "", "data_only", "ICS Excel", "Detail table; Excel-only"],
    ["Persona", "persona", "ax61", "Persona Velocity", "Line", 5, "chart_narrative", "ICS Excel", "Velocity curves + spend trajectory narrative"],
    ["Persona", "persona", "ax62", "Persona Cohort Trend", "Line", 5, "chart_narrative", "ICS Excel", "Cohort lines + persona evolution narrative"],
    # Referral Intelligence Engine
    ["Referral", "referral", "REF-1", "Top Referrers", "Horizontal Bar", 9, "screenshot", "ICS + Referral", "Full-width; ranked referrer list"],
    ["Referral", "referral", "REF-2", "Emerging Referrers", "Bar", 5, "chart_narrative", "ICS + Referral", "Chart + emerging referrer narrative"],
    ["Referral", "referral", "REF-3", "Dormant High-Value Referrers", "Table", "", "data_only", "ICS + Referral", "Action list; Excel-only"],
    ["Referral", "referral", "REF-4", "One-time vs Repeat Referrers", "Donut", 6, "comparison", "ICS + Referral", "Split: one-time left vs repeat right"],
    ["Referral", "referral", "REF-5", "Staff Multipliers", "Grouped Bar", 5, "chart_narrative", "ICS + Referral", "Chart + staff impact narrative"],
    ["Referral", "referral", "REF-6", "Branch Influence Density", "Heatmap", 9, "screenshot", "ICS + Referral", "Full-width heatmap; branches x influence"],
    ["Referral", "referral", "REF-7", "Code Health Report", "Bar", 5, "chart_narrative", "ICS + Referral", "Chart + code health narrative"],
    ["Referral", "referral", "REF-8", "Overview KPIs", "Multi-KPI", 7, "kpi_grid", "ICS + Referral", "6-cell KPI grid; runs last; aggregates"],
]


def _write_sheet(wb: Workbook, name: str, data: list[list]) -> None:
    ws = wb.create_sheet(title=name)

    # Header row
    for col_idx, header in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Data rows
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=10)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=col_idx == 9)

    # Auto-width columns
    for col_idx in range(1, len(COLUMNS) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(COLUMNS[col_idx - 1])
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 45)

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(data) + 1}"


LAYOUT_COLUMNS = [
    "Layout Index",
    "Layout Name",
    "Purpose",
    "Key Placeholders",
    "Used For",
    "Notes",
]

LAYOUT_DATA = [
    [
        0,
        "Cover / Intro - Slide 1",
        "Title slide with logo area, 3 text sections, KPI images",
        "ph[0] title, ph[32] subtitle, ph[14/33/34] 3 picture slots, "
        "ph[26-31] 3 paired header+body blocks, ph[19] body text",
        "Opening / title slide",
        "3 picture placeholders for logos/icons; 3 header+body pairs for KPI text",
    ],
    [
        1,
        "Cover / Intro - Slide 2",
        "Minimal cover, centered title only",
        "ph[0] title (6.3\" x 0.9\")",
        "Simple cover / closing",
        "No content areas; title is center-positioned",
    ],
    [
        2,
        "Divider - Slide 2",
        "Section divider with subtitle bar + large content area",
        "ph[0] title, ph[13] subtitle bar, ph[1] content (5.1\" x 4.0\")",
        "Section dividers between modules",
        "Left-aligned layout; content area can hold bullets or overview text",
    ],
    [
        3,
        "Divider - Slide 3",
        "Minimal section divider with subtitle bar",
        "ph[0] title, ph[13] subtitle bar",
        "Section dividers (minimal)",
        "Full-width title; no content area beyond subtitle",
    ],
    [
        4,
        "Analysis - Slide 1 (Chart + 2 Text)",
        "Chart on left, 2 header+body text sections on right",
        "ph[0] title, ph[13] subtitle, ph[14] picture (5.5\" x 3.4\" left), "
        "ph[26/19] top header+body, ph[27/28] bottom header+body",
        "screenshot_kpi slides (chart + KPI callouts)",
        "Right side has 2 independent text zones for insights/KPIs",
    ],
    [
        5,
        "Analysis - Slide 2 (Chart + 1 Text)",
        "Chart on left, single tall text section on right",
        "ph[0] title, ph[13] subtitle, ph[14] picture (5.5\" x 3.4\" left), "
        "ph[26] header, ph[19] body (5.7\" x 4.0\" right)",
        "screenshot_kpi slides (chart + narrative)",
        "Right body is 4\" tall -- good for longer narratives or bullet lists",
    ],
    [
        6,
        "Analysis - Slide 3 - Split",
        "Two equal content areas side-by-side",
        "ph[0] title, ph[13] subtitle, ph[14] left (5.8\" x 4.3\"), ph[1] right (5.8\" x 4.3\")",
        "multi_screenshot (2 charts side-by-side)",
        "Both zones are generic OBJECT placeholders; can hold images or text",
    ],
    [
        7,
        "Analysis - Slide 4 - Split",
        "6-cell KPI grid (3 header+body pairs per column)",
        "ph[0] title, ph[13] subtitle, ph[26-30] left 3 pairs, ph[32-36] right 3 pairs",
        "KPI summary / data grid slides",
        "12 text placeholders total; designed for structured KPI readouts",
    ],
    [
        8,
        "Analysis - Slide 5 - Thirds",
        "6-cell grid (3 columns x 2 rows of content blocks)",
        "ph[0] title, ph[13] subtitle, ph[1/15/16] top row, ph[17/19/20] bottom row",
        "Multi-chart grids, comparison panels",
        "Each cell is 3.6\" x 2.1\"; all OBJECT type (images or text)",
    ],
    [
        9,
        "Analysis - Slide 6 - Main",
        "Full-width single content area (the workhorse screenshot layout)",
        "ph[0] title, ph[13] subtitle, ph[1] content (11.7\" x 4.3\")",
        "screenshot (single full-width chart)",
        "Most common layout; used for any single chart/image",
    ],
    [
        10,
        "Analysis - Slide 7 - Stacked Two",
        "2 rows: each has text on left + 3 pictures on right",
        "ph[0] title, ph[15/35] row headers, ph[19/36] row body, "
        "ph[29-31] top 3 pics, ph[32-34] bottom 3 pics",
        "Before/after comparisons, multi-image galleries",
        "6 picture placeholders (2.1\" x 1.8\" each); good for small chart grids",
    ],
    [
        11,
        "Analysis - Slide 8 - Blank",
        "Blank slide with right-aligned title",
        "ph[0] title (6.4\" x 1.6\", right side)",
        "Custom-built slides, freeform content",
        "No content placeholders; shapes must be added programmatically",
    ],
    [
        12,
        "Analysis - Slide 11 - Header1",
        "Pure background / decorative header",
        "(no placeholders)",
        "Visual separator, background-only",
        "No usable placeholders; purely decorative",
    ],
    [
        13,
        "Analysis - Slide 11 - Header2",
        "Background header with footer/slide number",
        "ph[11] footer, ph[12] slide number",
        "Mailer summaries, wide charts (custom positioning)",
        "No title/content placeholders; content placed via freeform shapes",
    ],
]


SLIDE_TYPE_COLUMNS = [
    "Slide Type",
    "Layout Index",
    "Layout Name",
    "When To Use",
    "Content Zones",
    "Examples",
]

SLIDE_TYPE_DATA = [
    [
        "chart_narrative",
        5,
        "Analysis - Slide 2 (Chart + 1 Text)",
        "Any chart that needs explanatory text -- trend lines, donuts, scatters, funnels, waterfalls",
        "Chart (5.5\" x 3.4\" left) + header + tall body (5.7\" x 4.0\" right)",
        "DCTR-1 trend, A7.8 funnel, A11.1 waterfall, ax08 source donut, ax29 milestones",
    ],
    [
        "chart_kpi",
        4,
        "Analysis - Slide 1 (Chart + 2 Text)",
        "Charts with 2 KPI callout zones -- rates, snapshots, impact metrics with headline numbers",
        "Chart (5.5\" x 3.4\" left) + top header+body + bottom header+body (right)",
        "DCTR-3 snapshot, A8.3 Reg E snapshot, A9.1 attrition rate, A9.9-A9.11 impact KPIs",
    ],
    [
        "comparison",
        6,
        "Analysis - Slide 3 - Split",
        "Natural A-vs-B content -- Personal vs Business, Open vs Closed, DM vs REF, rank vs change",
        "Left content (5.8\" x 4.3\") + right content (5.8\" x 4.3\")",
        "DCTR-4/5 personal vs business, A9.13 ARS vs non-ARS, ax14 age comparison, REF-4 one-time vs repeat",
    ],
    [
        "multi_screenshot",
        6,
        "Analysis - Slide 3 - Split",
        "Merged chart pairs -- two related charts that tell one story side by side",
        "Left content (5.8\" x 4.3\") + right content (5.8\" x 4.3\")",
        "A7.6a trajectory+segments, A7.7 historical vs TTM, A8.5 opportunity, A9.3 open vs closed",
    ],
    [
        "kpi_grid",
        7,
        "Analysis - Slide 4 - Split",
        "Pure KPI summary slides -- structured readouts with 6 metric cells",
        "6 cells: 3 header+body pairs per column (left 3, right 3)",
        "A7.10c branch KPIs, A8.12 Reg E summary, ax01/02 ICS summary, ax22 activity, REF-8 overview",
    ],
    [
        "grid_thirds",
        8,
        "Analysis - Slide 5 - Thirds",
        "Multi-panel charts -- 3+ small charts or comparison panels in a grid",
        "6 cells: 3 columns x 2 rows (each 3.6\" x 2.1\")",
        "MCC 3-panel comparison, portfolio scorecard bullet gauges",
    ],
    [
        "screenshot",
        9,
        "Analysis - Slide 6 - Main",
        "Data-dense charts that need full width -- heatmaps, long merchant lists, many-branch bars",
        "Full-width content (11.7\" x 4.3\")",
        "DCTR-8 summary, DCTR-9 heatmap, top-25 merchants, branch rank bars, cohort heatmaps",
    ],
    [
        "section_text",
        2,
        "Divider - Slide 2",
        "Text-only slides -- conclusions, section intros with explanatory content",
        "Title + subtitle bar + large text content area (5.1\" x 4.0\")",
        "S6-S8 conclusions, section overview text",
    ],
    [
        "wide_custom",
        13,
        "Analysis - Slide 11 - Header2",
        "Dense matrices needing freeform positioning -- wide heatmaps, custom mailer layouts",
        "No content placeholders; shapes added programmatically",
        "A7.10a branch matrix, A8.4a branch Reg E matrix, mailer monthly summaries",
    ],
    [
        "data_only",
        "",
        "(no PPTX slide)",
        "Excel-only tables and data engine steps -- no visual chart, no slide generated",
        "N/A -- data appears in Excel report only",
        "MCC tables, monthly movers, competitor detection, dormant lists, overview tables",
    ],
]


def _write_slide_type_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet(title="Slide Types")

    for col_idx, header in enumerate(SLIDE_TYPE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for row_idx, row_data in enumerate(SLIDE_TYPE_DATA, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=10)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=col_idx >= 4)

    widths = [18, 14, 36, 70, 55, 70]
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(SLIDE_TYPE_COLUMNS))}{len(SLIDE_TYPE_DATA) + 1}"


def _write_layout_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet(title="Template Layouts")

    # Header row
    for col_idx, header in enumerate(LAYOUT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Data rows
    for row_idx, row_data in enumerate(LAYOUT_DATA, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=10)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=col_idx >= 4)

    # Column widths
    widths = [14, 36, 52, 70, 42, 60]
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(LAYOUT_COLUMNS))}{len(LAYOUT_DATA) + 1}"


PATH_COLUMNS = ["Category", "Package", "File Path", "Purpose"]

# Relative to packages/ for brevity -- full prefix is packages/<pkg>/src/<pkg>/
PATH_DATA = [
    # --- Config ---
    ["Config", "ars_analysis", "ars_analysis/config.py", "PathsConfig, CSMSourcesConfig, PipelineConfig, ARSSettings (loads ars_config.json)"],
    ["Config", "shared", "shared/config.py", "PlatformConfig (base_output_dir, m_drive_path, chart_theme, template_pptx, pipelines)"],
    ["Config", "txn_analysis", "txn_analysis/settings.py", "Settings (data_file, odd_file, output_dir, charts, outputs, segments)"],
    ["Config", "ics_toolkit", "ics_toolkit/settings.py", "AppendSettings, AnalysisSettings, ReferralSettings"],
    ["Config", "root", "config/platform.yaml", "Top-level platform config (per-pipeline settings)"],
    ["Config", "root", "config/clients_config.json", "Master client config (DataStartDate, ICRate, NSF_OD_Fee, BranchMapping, etc.)"],
    ["Config", "root", "config/benchmarks.json", "Industry benchmarks (debit penetration, active card rate, avg spend)"],
    # --- Data Loaders ---
    ["Data Loader", "shared", "shared/data_loader.py", "load_oddd(), load_tran(), load_odd(), _read_file() (auto-detect xlsx/xls/csv)"],
    ["Data Loader", "ars_analysis", "ars_analysis/pipeline/steps/load.py", "Load ODD file for ARS pipeline"],
    ["Data Loader", "txn_analysis", "txn_analysis/data_loader.py", "load_data(), load_odd(), merge_odd(), _apply_merchant_consolidation(), _derive_year_month()"],
    ["Data Loader", "ics_toolkit", "ics_toolkit/analysis/data_loader.py", "load_data(), _normalize_strings(), _parse_dates(), _coerce_numerics(), discover L12M cols"],
    ["Data Loader", "ics_toolkit", "ics_toolkit/referral/data_loader.py", "Referral-specific data loading and normalization"],
    # --- Retrieve / Format ---
    ["Retrieve/Format", "ars_analysis", "ars_analysis/pipeline/steps/retrieve.py", "retrieve_all() -- copy ODD from CSM M: drive sources to retrieve_dir/CSM/YYYY.MM/ClientID/"],
    ["Retrieve/Format", "shared", "shared/format_odd.py", "format_odd() 7-step pipeline (PYTD drop, totals, PIN+Sig, age, response, segmentation)"],
    ["Retrieve/Format", "ars_analysis", "ars_analysis/pipeline/steps/format.py", "format_all() -- batch format from retrieve_dir to watch_root"],
    ["Retrieve/Format", "ars_analysis", "ars_analysis/pipeline/steps/scan.py", "scan_ready_files(), _pick_best_file() (prefers *-formatted.xlsx), available_months/csms"],
    # --- Column Mapping ---
    ["Column Mapping", "txn_analysis", "txn_analysis/column_map.py", "COLUMN_ALIASES (50+ header variations), resolve_columns(), REQUIRED/OPTIONAL_COLUMNS"],
    ["Column Mapping", "ics_toolkit", "ics_toolkit/analysis/column_map.py", "resolve_columns(), validate_columns(), discover_l12m_columns()"],
    ["Column Mapping", "ics_toolkit", "ics_toolkit/referral/column_map.py", "Referral column normalization"],
    # --- Reference Data ---
    ["Reference Data", "txn_analysis", "txn_analysis/competitor_patterns.py", "COMPETITOR_MERCHANTS (7 categories, 3-tier matching), FINANCIAL_MCC_CODES, classify_merchant()"],
    ["Reference Data", "txn_analysis", "txn_analysis/segments.py", "extract_responder_accounts(), extract_ics_accounts(), build_segment_filters()"],
    ["Reference Data", "txn_analysis", "txn_analysis/formatting.py", "excel_number_format(), is_percentage_column(), is_grand_total_row()"],
    ["Reference Data", "ics_toolkit", "ics_toolkit/referral/code_decoder.py", "Referral code decoding and classification"],
    ["Reference Data", "ics_toolkit", "ics_toolkit/referral/scoring.py", "Referral scoring weights and calculations"],
    # --- Deck Builders ---
    ["Deck Builder", "shared", "shared/deck/__init__.py", "Re-exports: DeckBuilder, SlideContent, build_deck_from_results"],
    ["Deck Builder", "shared", "shared/deck/engine.py", "DeckBuilder class + SlideContent dataclass (extracted reusable engine)"],
    ["Deck Builder", "shared", "shared/deck/universal.py", "build_deck_from_results() -- universal builder for TXN/ICS/Attrition"],
    ["Deck Builder", "ars_analysis", "ars_analysis/output/deck_builder.py", "ARS-specific deck builder (500+ lines, section ordering, mailer logic)"],
    ["Deck Builder", "ics_toolkit", "ics_toolkit/analysis/exports/deck_builder.py", "ICS analysis deck builder"],
    ["Deck Builder", "ics_toolkit", "ics_toolkit/analysis/exports/pptx.py", "ICS PPTX export helpers"],
    ["Deck Builder", "ics_toolkit", "ics_toolkit/analysis/exports/kpi_slides.py", "ICS KPI slide builder"],
    ["Deck Builder", "ics_toolkit", "ics_toolkit/referral/exports/pptx.py", "ICS referral deck builder"],
    ["Deck Builder", "txn_analysis", "txn_analysis/exports/pptx_report.py", "TXN PPTX report (charts-only deck)"],
    # --- Templates ---
    ["Template", "shared", "shared/deck/template/Template12.25.pptx", "Shared PPTX template (14 layouts) -- canonical copy"],
    ["Template", "ars_analysis", "ars_analysis/output/template/Template12.25.pptx", "ARS copy of template"],
    ["Template", "ics_toolkit", "ics_toolkit/templates/Template12.25.pptx", "ICS copy of template"],
    # --- Excel Exports ---
    ["Excel Export", "shared", "shared/excel.py", "Shared Excel helpers (formatting, styles)"],
    ["Excel Export", "ars_analysis", "ars_analysis/output/excel_formatter.py", "ARS Excel report (Summary sheet, KPI extraction, per-section tabs)"],
    ["Excel Export", "ics_toolkit", "ics_toolkit/analysis/exports/excel.py", "ICS analysis Excel report"],
    ["Excel Export", "ics_toolkit", "ics_toolkit/referral/exports/excel.py", "ICS referral Excel report"],
    ["Excel Export", "txn_analysis", "txn_analysis/exports/excel_report.py", "TXN Excel report (cover, TOC, per-analysis sheets, chart embeds)"],
    # --- Chart Libraries ---
    ["Charts", "shared", "shared/charts.py", "Shared chart utilities (chart_figure context manager, style isolation)"],
    ["Charts", "ars_analysis", "ars_analysis/charts/__init__.py", "ARS chart registry + style"],
    ["Charts", "ars_analysis", "ars_analysis/charts/guards.py", "ARS matplotlib leak guards"],
    ["Charts", "txn_analysis", "txn_analysis/charts/__init__.py", "TXN chart registry (CHART_REGISTRY dict), create_charts(), render_chart_png()"],
    ["Charts", "txn_analysis", "txn_analysis/charts/builders.py", "TXN chart primitives (lollipop, donut, heatmap, grouped_bar, etc.)"],
    ["Charts", "txn_analysis", "txn_analysis/charts/theme.py", "TXN chart theme + add_source_footer()"],
    ["Charts", "txn_analysis", "txn_analysis/charts/{overall,personal,business,mcc,competitor,trends,recurring,scorecard}.py", "TXN per-section chart functions"],
    ["Charts", "ics_toolkit", "ics_toolkit/analysis/charts/__init__.py", "ICS chart registry"],
    ["Charts", "ics_toolkit", "ics_toolkit/analysis/charts/renderer.py", "ICS chart renderer (PNG export)"],
    ["Charts", "ics_toolkit", "ics_toolkit/analysis/charts/{summary,source,dm_source,activity,demographics,cohort,performance,portfolio,persona,strategic}.py", "ICS per-section chart functions"],
    ["Charts", "ics_toolkit", "ics_toolkit/referral/charts/{top_referrers,branch_density,code_health,staff_multipliers,emerging_referrers}.py", "ICS referral chart functions"],
    # --- Analysis Modules ---
    ["Analysis", "ars_analysis", "ars_analysis/analytics/registry.py", "ARS module registry (ABC + @register)"],
    ["Analysis", "ars_analysis", "ars_analysis/analytics/overview/{stat_codes,product_codes,eligibility}.py", "Overview: A1, A1b, A3"],
    ["Analysis", "ars_analysis", "ars_analysis/analytics/dctr/{penetration,trends,branches,funnel,overlays,_helpers}.py", "DCTR: DCTR-1 to 16, A7.x (26 slides)"],
    ["Analysis", "ars_analysis", "ars_analysis/analytics/rege/{status,branches,dimensions,_helpers}.py", "Reg E: A8.x (13 slides)"],
    ["Analysis", "ars_analysis", "ars_analysis/analytics/attrition/{rates,dimensions,impact,_helpers}.py", "Attrition: A9.x (13 slides)"],
    ["Analysis", "ars_analysis", "ars_analysis/analytics/value/analysis.py", "Value: A11.1, A11.2"],
    ["Analysis", "ars_analysis", "ars_analysis/analytics/mailer/{reach,response,impact,insights,cohort,_helpers}.py", "Mailer: A12-A15 (dynamic per month)"],
    ["Analysis", "ars_analysis", "ars_analysis/analytics/insights/{synthesis,conclusions,effectiveness,dormant,branch_scorecard,_data}.py", "Insights: S1-S8"],
    ["Analysis", "ics_toolkit", "ics_toolkit/analysis/analyses/{summary,source,dm_source,ref_source,demographics,activity,cohort,performance,portfolio,persona,strategic}.py", "ICS: 82 chart slides across 12 sections"],
    ["Analysis", "ics_toolkit", "ics_toolkit/referral/analyses/{overview,top_referrers,branch_density,code_health,emerging_referrers,dormant_referrers,staff_multipliers,onetime_vs_repeat}.py", "Referral: REF-1 to REF-8"],
    ["Analysis", "txn_analysis", "txn_analysis/analyses/{overall,personal,business,mcc,financial_services,interchange,member_segments}.py", "TXN: M1-M4, M7-M8, M10"],
    ["Analysis", "txn_analysis", "txn_analysis/analyses/{competitor_detect,competitor_metrics,competitor_segment,competitor_threat}.py", "TXN: M6 competitor"],
    ["Analysis", "txn_analysis", "txn_analysis/analyses/{trends_rank,trends_growth,trends_consistency,trends_cohort,trends_movers}.py", "TXN: M5 trends"],
    ["Analysis", "txn_analysis", "txn_analysis/analyses/{recurring,scorecard,spending_behavior,time_patterns}.py", "TXN: M9, M15, behavioral"],
    # --- Pipelines & Orchestration ---
    ["Pipeline", "platform_app", "platform_app/orchestrator.py", "Central orchestrator + _ensure_deck() fallback"],
    ["Pipeline", "ars_analysis", "ars_analysis/pipeline/runner.py", "ARS pipeline runner (load -> subset -> analyze -> generate)"],
    ["Pipeline", "ars_analysis", "ars_analysis/pipeline/context.py", "ARS PipelineContext dataclass"],
    ["Pipeline", "ars_analysis", "ars_analysis/pipeline/batch.py", "ARS batch runner (run_batch for 300+ clients)"],
    ["Pipeline", "ars_analysis", "ars_analysis/runner.py", "ARS CLI entry point"],
    ["Pipeline", "ics_toolkit", "ics_toolkit/analysis/pipeline.py", "ICS analysis pipeline"],
    ["Pipeline", "ics_toolkit", "ics_toolkit/referral/pipeline.py", "ICS referral pipeline"],
    ["Pipeline", "ics_toolkit", "ics_toolkit/runner.py", "ICS CLI entry point"],
    ["Pipeline", "txn_analysis", "txn_analysis/pipeline.py", "TXN pipeline (load -> analyze -> chart -> export)"],
    ["Pipeline", "txn_analysis", "txn_analysis/runner.py", "TXN CLI entry point"],
    # --- UI Pages ---
    ["UI", "platform_app", "platform_app/pages/pipeline_ars.py", "Streamlit ARS pipeline page"],
    ["UI", "platform_app", "platform_app/pages/pipeline_txn.py", "Streamlit TXN pipeline page"],
    ["UI", "platform_app", "platform_app/pages/pipeline_ics.py", "Streamlit ICS pipeline page"],
    ["UI", "platform_app", "platform_app/pages/pipeline_attrition.py", "Streamlit Attrition pipeline page"],
    # --- Scripts ---
    ["Script", "root", "run.bat", "4-step pipeline: retrieve -> format -> batch -> streamlit"],
    ["Script", "root", "run_batch.bat", "Headless batch runner"],
    ["Script", "root", "dashboard.bat", "Launch Streamlit UI only"],
    ["Script", "root", "setup.bat", "Environment setup (uv install)"],
    # --- Test Data ---
    ["Test Data", "root", "tests/e2e_data/1200_Test CU_2026.02.xlsx", "Synthetic ARS ODD (20 accounts)"],
    ["Test Data", "root", "tests/e2e_data/8888_transactions.csv", "Synthetic TXN (35 transactions)"],
    ["Test Data", "root", "tests/e2e_data/9999_ICS_2026.01.xlsx", "Synthetic ICS (80 accounts)"],
    ["Test Data", "root", "tests/e2e_data/generate_fixtures.py", "E2E fixture generator script"],
    # --- Catalog ---
    ["Catalog", "root", "docs/slide_catalog.py", "This file -- generates the catalog Excel"],
    ["Catalog", "root", "docs/slide_catalog.xlsx", "Generated workbook (all tabs)"],
]

DATA_SOURCE_COLUMNS = ["Pipeline", "Data Type", "Source Location", "Format", "Key Columns", "Notes"]

DATA_SOURCE_DATA = [
    # ARS
    ["ARS", "ODD (raw)", r"M:\CSM-Source\Folder\ClientID_ODDD*.xlsx", "Excel (.xlsx)",
     "Account #, Date Opened, Date Closed, Stat Code, Prod Code, Branch, DOB, Balance fields",
     "Copied by retrieve_all() to retrieve_dir/CSM/YYYY.MM/ClientID/"],
    ["ARS", "ODD (formatted)", r"watch_root\CSM\YYYY.MM\ClientID\*-formatted.xlsx", "Excel (.xlsx)",
     "All raw cols + Total Spend, Total Swipes, Avg Monthly, Account Age, Holder Age, Response Grouping, Segmentation",
     "Created by format_odd() 7-step pipeline; scan_ready_files() prefers these"],
    ["ARS", "Client Config", r"M:\ARS\Config\clients_config.json  OR  config/clients_config.json", "JSON",
     "ClientID, DataStartDate, ICRate, NSF_OD_Fee, EligibleStatusCodes, BranchMapping",
     "Per-client overrides; falls back to defaults"],
    ["ARS", "ARS Config", "ars_config.json (project root or ARS_CONFIG_PATH env)", "JSON",
     "paths (ars_base, watch_root, retrieve_dir), csm_sources, pipeline settings",
     "Loaded by ARSSettings; env vars with ARS_ prefix override"],
    # TXN
    ["TXN", "Transaction CSV", "User-provided or auto-discovered near ODD file", "CSV/TXT (tab-delimited)",
     "merchant_name, amount, primary_account_num, transaction_date, mcc_code, business_flag, year_month",
     "50+ column aliases auto-resolved by resolve_columns(); 13-col standard layout"],
    ["TXN", "ODD (for segments)", "Via --odd CLI flag or Settings.odd_file", "Excel (.xlsx)",
     "Account #, Segmentation cols (MmmYY Seg), ICS Account, generation, tenure, balance_tier",
     "Optional; enables M11 demographics, segment filters (ARS responders, ICS accounts)"],
    ["TXN", "Transaction Dir", "Settings.transaction_dir (year-folder layout)", "Directory of CSVs",
     "YYYY/MM/*.csv files; auto-selects most recent 12 months",
     "Alternative to single file; V4 multi-month mode"],
    # ICS
    ["ICS", "ICS Excel", "User-provided or M: drive", "Excel (.xlsx)",
     "Account #, ICS Account (Yes/No), ICS Source (REF/DM/Both), Stat Code, Prod Code, Branch, Balance, Date Opened/Closed",
     "Core input; L12M monthly columns auto-discovered (MmmYY pattern)"],
    ["ICS", "Referral Data", "Same ICS Excel or separate referral file", "Excel (.xlsx)",
     "Referring Account, Referred Account, Referral Date, Referral Code, Branch",
     "Used by referral pipeline (REF-1 to REF-8); requires code_decoder for code classification"],
    ["ICS", "Revenue Workbook", "ICS directory or config path", "Excel (.xlsx)",
     "Account #, Monthly Interchange, Annual Revenue",
     "Optional; enables revenue impact analyses (ax39, ax65, ax66)"],
    ["ICS", "Benchmarks", "config/benchmarks.json", "JSON",
     "debit_penetration_rate, active_card_rate, avg_annual_spend_per_card, direct_mail_response_rate",
     "Industry benchmarks for comparison analyses"],
    # Shared
    ["All", "PPTX Template", "packages/shared/src/shared/deck/template/Template12.25.pptx", "PowerPoint (.pptx)",
     "14 layouts (cover, divider, chart+text, split, grid, blank, etc.)",
     "Identical copies in ars_analysis and ics_toolkit packages"],
    ["All", "Platform Config", "config/platform.yaml", "YAML",
     "base_output_dir, m_drive_path, chart_theme, per-pipeline enabled/settings",
     "Top-level orchestrator config"],
]

M_DRIVE_COLUMNS = ["Path", "Purpose", "Read/Write", "Used By"]

M_DRIVE_DATA = [
    ["M:\\ARS\\", "ARS base directory", "R/W", "All ARS operations"],
    ["M:\\ARS\\Ready for Analysis\\CSM\\YYYY.MM\\ClientID\\", "Formatted ODD files (watch_root)", "R/W", "scan, analyze, batch"],
    ["M:\\ARS\\Incoming\\ODDD Files\\CSM\\YYYY.MM\\ClientID\\", "Raw ODD files (retrieve_dir)", "R/W", "retrieve_all()"],
    ["M:\\ARS\\Presentations\\Presentation Excels\\", "Output Excel reports", "Write", "excel_formatter, batch"],
    ["M:\\ARS\\Presentations\\Presentation Excels\\Archive\\", "Archived previous reports", "Write", "batch (auto-archive)"],
    ["M:\\ARS\\Presentations\\Template12.25.pptx", "PPTX template (runtime)", "Read", "deck_builder"],
    ["M:\\ARS\\Scripts\\Config\\", "Runtime config directory", "Read", "ARSSettings"],
    ["M:\\ARS\\Config\\clients_config.json", "Master client config", "Read", "Pipeline context, per-client overrides"],
    ["M:\\ARS\\Logs\\", "Pipeline log files", "Write", "All pipelines"],
    ["M:\\CSM-Source\\{CSMName}\\", "CSM source directories (per ars_config.json csm_sources)", "Read", "retrieve_all()"],
    ["M:\\ICS\\Config\\clients_config.json", "ICS client config fallback", "Read", "ICS pipeline"],
]


# ── Critical Field Names ──────────────────────────────────────────────

FIELD_COLUMNS = ["Pipeline", "Category", "Field Name", "Aliases", "Required", "Type", "Source", "Notes"]

FIELD_DATA = [
    # ── ODD: Raw input columns ──
    ["ODD", "Core", "Acct Number", "", "Yes", "str", "Raw Excel", "Primary key for cross-pipeline merges"],
    ["ODD", "Core", "Stat Code", "Status Code, StatCode, Stat_Code, Account Status", "Yes", "str", "Raw Excel", "Account status (O=Open, C=Closed)"],
    ["ODD", "Core", "Product Code", "Prod Code, ProdCode, Prod_Code", "Yes", "str", "Raw Excel", "Product type identifier"],
    ["ODD", "Core", "Date Opened", "DateOpened, Date_Opened, Open Date", "Yes", "datetime", "Raw Excel", "Account open date"],
    ["ODD", "Core", "Date Closed", "DateClosed, Date_Closed, Close Date", "No", "datetime", "Raw Excel", "Account close date (NaT if open)"],
    ["ODD", "Core", "Avg Bal", "Balance, Current Balance, Cur Bal, AvgBal, Avg_Bal, Average Balance", "Yes", "float", "Raw Excel", "Average balance"],
    ["ODD", "Core", "Branch", "", "No", "str", "Raw Excel", "Branch identifier"],
    ["ODD", "Core", "Business?", "Business, BusinessFlag, Business Flag", "No", "str", "Raw Excel", "Personal vs business flag"],
    ["ODD", "Core", "Debit?", "Debit, DC Indicator, DC_Indicator, DebitCard, Debit Card", "No", "str", "Raw Excel", "Debit card indicator (YES/Y/D/DC/DEBIT)"],
    ["ODD", "Core", "DOB", "", "No", "datetime", "Raw Excel", "Date of birth (used to compute age)"],
    ["ODD", "Core", "Mailable?", "", "No", "str", "Raw Excel", "Eligibility flag for mailers"],
    # ── ODD: Monthly time-series ──
    ["ODD", "Monthly", "MmmYY PIN $", "e.g. Jan25 PIN $", "No", "float", "Raw Excel", "PIN debit dollar amount per month"],
    ["ODD", "Monthly", "MmmYY Sig $", "e.g. Jan25 Sig $", "No", "float", "Raw Excel", "Signature debit dollar amount per month"],
    ["ODD", "Monthly", "MmmYY PIN #", "e.g. Jan25 PIN #", "No", "int", "Raw Excel", "PIN debit transaction count per month"],
    ["ODD", "Monthly", "MmmYY Sig #", "e.g. Jan25 Sig #", "No", "int", "Raw Excel", "Signature debit transaction count per month"],
    ["ODD", "Monthly", "MmmYY MTD", "e.g. Jan25 MTD", "No", "int", "Raw Excel", "Month-to-date items"],
    ["ODD", "Monthly", "MmmYY OD Limit", "e.g. Jan25 OD Limit", "No", "float", "Raw Excel", "Overdraft limit per month"],
    ["ODD", "Monthly", "MmmYY Reg E Code", "e.g. Jan25 Reg E Code", "No", "str", "Raw Excel", "Reg E status code per month"],
    ["ODD", "Monthly", "MmmYY Reg E Desc", "e.g. Jan25 Reg E Desc", "No", "str", "Raw Excel", "Reg E description per month"],
    ["ODD", "Monthly", "MmmYY Mail", "e.g. Jan25 Mail", "No", "str", "Raw Excel", "Mailer segment code per month"],
    ["ODD", "Monthly", "MmmYY Resp", "e.g. Jan25 Resp", "No", "str", "Raw Excel", "Response segment code per month"],
    # ── ODD: Derived by format_odd() ──
    ["ODD", "Derived", "MmmYY Spend", "e.g. Jan25 Spend", "N/A", "float", "format_odd", "PIN $ + Sig $ combined per month"],
    ["ODD", "Derived", "MmmYY Swipes", "e.g. Jan25 Swipes", "N/A", "int", "format_odd", "PIN # + Sig # combined per month"],
    ["ODD", "Derived", "Total Spend", "", "N/A", "float", "format_odd", "Sum of all monthly Spend"],
    ["ODD", "Derived", "Total Swipes", "", "N/A", "int", "format_odd", "Sum of all monthly Swipes"],
    ["ODD", "Derived", "Total Items", "", "N/A", "int", "format_odd", "Sum of all monthly MTD"],
    ["ODD", "Derived", "last 3-mon spend", "", "N/A", "float", "format_odd", "Last 3 months cumulative spend"],
    ["ODD", "Derived", "last 3-mon swipes", "", "N/A", "int", "format_odd", "Last 3 months cumulative swipes"],
    ["ODD", "Derived", "Last 3-mon Items", "", "N/A", "int", "format_odd", "Last 3 months cumulative items"],
    ["ODD", "Derived", "last 12-mon spend", "", "N/A", "float", "format_odd", "Last 12 months cumulative spend"],
    ["ODD", "Derived", "last 12-mon swipes", "", "N/A", "int", "format_odd", "Last 12 months cumulative swipes"],
    ["ODD", "Derived", "Last 12-mon Items", "", "N/A", "int", "format_odd", "Last 12 months cumulative items"],
    ["ODD", "Derived", "MonthlySpend12", "", "N/A", "float", "format_odd", "12-month avg monthly spend"],
    ["ODD", "Derived", "MonthlySwipes12", "", "N/A", "float", "format_odd", "12-month avg monthly swipes"],
    ["ODD", "Derived", "MonthlyItems12", "", "N/A", "float", "format_odd", "12-month avg monthly items"],
    ["ODD", "Derived", "MonthlySpend3", "", "N/A", "float", "format_odd", "3-month avg monthly spend"],
    ["ODD", "Derived", "MonthlySwipes3", "", "N/A", "float", "format_odd", "3-month avg monthly swipes"],
    ["ODD", "Derived", "MonthlyItems3", "", "N/A", "float", "format_odd", "3-month avg monthly items"],
    ["ODD", "Derived", "SwipeCat12", "", "N/A", "str", "format_odd", "12-month swipe tier (Non-user, 1-5, 6-10, 11-20, 21-40, 41+)"],
    ["ODD", "Derived", "SwipeCat3", "", "N/A", "str", "format_odd", "3-month swipe tier"],
    ["ODD", "Derived", "Account Holder Age", "", "N/A", "float", "format_odd", "Years from DOB to report date"],
    ["ODD", "Derived", "Account Age", "", "N/A", "float", "format_odd", "Years from Date Opened to close/report date"],
    ["ODD", "Derived", "# of Offers", "", "N/A", "int", "format_odd", "Count of non-null Mail columns"],
    ["ODD", "Derived", "# of Responses", "", "N/A", "int", "format_odd", "Count of non-null Resp columns (excl NU 1-4)"],
    ["ODD", "Derived", "Response Grouping", "", "N/A", "str", "format_odd", "No Offer / Non-Responder / SO-SR / MO-SR / MR"],
    ["ODD", "Derived", "MmmYY Segmentation", "e.g. Jan25 Segmentation", "N/A", "str", "format_odd", "Control / Non-Responder / Responder per month"],
    ["ODD", "Derived", "ICS Account", "", "N/A", "str", "ICS append", "Yes / No (appended by ICS pipeline)"],
    ["ODD", "Derived", "ICS Source", "", "N/A", "str", "ICS append", "REF / DM / Both / empty"],
    # ── TXN: Required columns ──
    ["TXN", "Required", "merchant_name", "merchantname, merchant, merch_name, description, payee, vendor, ...", "Yes", "str", "CSV", "Merchant/payee name (12 aliases)"],
    ["TXN", "Required", "amount", "transaction_amount, txn_amount, amt, debit_amount, purchase_amount, ...", "Yes", "float", "CSV", "Transaction dollar amount (11 aliases)"],
    ["TXN", "Required", "primary_account_num", "account_number, acct_num, card_number, pan, member_number, ...", "Yes", "str", "CSV", "Account identifier (18 aliases)"],
    ["TXN", "Required", "transaction_date", "trans_date, txn_date, date, posting_date, settlement_date, ...", "Yes", "datetime", "CSV", "Transaction date (12 aliases)"],
    # ── TXN: Optional columns ──
    ["TXN", "Optional", "mcc_code", "mcccode, mcc, merchant_category_code, sic_code, sic", "No", "str", "CSV", "Merchant Category Code (7 aliases)"],
    ["TXN", "Optional", "business_flag", "businessflag, business, is_business, account_type", "No", "str", "CSV", "Business account indicator (6 aliases)"],
    ["TXN", "Optional", "year_month", "yearmonth, ym", "No", "str", "CSV", "Pre-computed YYYY-MM period"],
    # ── TXN: Tab-delimited positional (V4 format) ──
    ["TXN", "V4 Positional", "transaction_type", "", "No", "str", "Tab file col 3", "Transaction type code"],
    ["TXN", "V4 Positional", "terminal_location_1", "", "No", "str", "Tab file col 7", "Terminal location line 1"],
    ["TXN", "V4 Positional", "terminal_location_2", "", "No", "str", "Tab file col 8", "Terminal location line 2"],
    ["TXN", "V4 Positional", "terminal_id", "", "No", "str", "Tab file col 9", "Terminal identifier"],
    ["TXN", "V4 Positional", "merchant_id", "", "No", "str", "Tab file col 10", "Merchant identifier"],
    ["TXN", "V4 Positional", "institution", "", "No", "str", "Tab file col 11", "Issuing institution"],
    ["TXN", "V4 Positional", "card_present", "", "No", "str", "Tab file col 12", "Card-present indicator"],
    ["TXN", "V4 Positional", "transaction_code", "", "No", "str", "Tab file col 13", "Transaction code"],
    # ── TXN: Derived columns ──
    ["TXN", "Derived", "merchant_consolidated", "", "N/A", "str", "data_loader", "Standardized merchant name via rules"],
    ["TXN", "Derived", "year_month", "(if not in CSV)", "N/A", "str", "data_loader", "YYYY-MM derived from transaction_date"],
    ["TXN", "Derived", "is_partial_month", "", "N/A", "bool", "data_loader", "True if last month has <90% day coverage"],
    ["TXN", "Derived", "source_file", "", "N/A", "str", "data_loader", "Originating CSV filename"],
    ["TXN", "Derived", "generation", "", "N/A", "str", "ODD merge", "Gen Z / Millennial / Gen X / Boomer / Silent"],
    ["TXN", "Derived", "balance_tier", "", "N/A", "str", "ODD merge", "Low / Medium / High / Very High"],
    ["TXN", "Derived", "tenure_years", "", "N/A", "float", "ODD merge", "Years since account opened"],
    # ── ICS Analysis: Required columns ──
    ["ICS", "Required", "ICS Account", "ICS Accounts, Ics Account, ICS_Account, IcsAccount", "Yes", "str", "Excel", "ICS account flag (Yes/No)"],
    ["ICS", "Required", "Stat Code", "StatCode, Stat_Code, Status Code", "Yes", "str", "Excel", "Account status code"],
    ["ICS", "Required", "Debit?", "Debit, DebitCard, Debit Card", "Yes", "str", "Excel", "Debit card indicator"],
    ["ICS", "Required", "Business?", "Business, BusinessFlag, Business Flag", "Yes", "str", "Excel", "Personal vs business flag"],
    ["ICS", "Required", "Date Opened", "DateOpened, Date_Opened, Open Date", "Yes", "datetime", "Excel", "Account open date"],
    ["ICS", "Required", "Prod Code", "ProdCode, Prod_Code, Product Code", "Yes", "str", "Excel", "Product code"],
    ["ICS", "Required", "Branch", "", "Yes", "str", "Excel", "Branch identifier"],
    ["ICS", "Required", "Source", "", "Yes", "str", "Excel", "ICS source (REF/DM/Both)"],
    ["ICS", "Required", "Curr Bal", "CurrBal, Curr_Bal, Current Balance", "Yes", "float", "Excel", "Current balance"],
    # ── ICS Analysis: Optional columns ──
    ["ICS", "Optional", "Date Closed", "DateClosed, Date_Closed, Close Date", "No", "datetime", "Excel", "Account close date"],
    ["ICS", "Optional", "Avg Bal", "AvgBal, Avg_Bal, Average Balance, AvgColBal", "No", "float", "Excel", "Average balance"],
    # ── ICS Analysis: L12M dynamic columns ──
    ["ICS", "L12M", "MmmYY Swipes", "regex: ^[A-Z][a-z]{2}\\d{2} Swipes$", "No", "int", "Excel", "Monthly swipe count (e.g. Feb24 Swipes)"],
    ["ICS", "L12M", "MmmYY Spend", "regex: ^[A-Z][a-z]{2}\\d{2} Spend$", "No", "float", "Excel", "Monthly spend (e.g. Feb24 Spend)"],
    # ── ICS Analysis: Derived columns ──
    ["ICS", "Derived", "Total L12M Swipes", "", "N/A", "int", "utils.py", "Sum of all MmmYY Swipes columns"],
    ["ICS", "Derived", "Total L12M Spend", "", "N/A", "float", "utils.py", "Sum of all MmmYY Spend columns"],
    ["ICS", "Derived", "Active in L12M", "", "N/A", "bool", "utils.py", "True if Total L12M Swipes > 0"],
    ["ICS", "Derived", "Opening Month", "", "N/A", "str", "utils.py", "YYYY-MM from Date Opened"],
    ["ICS", "Derived", "Account Age Days", "", "N/A", "int", "utils.py", "Days since Date Opened"],
    ["ICS", "Derived", "Balance Tier", "", "N/A", "str", "utils.py", "Binned from Curr Bal"],
    ["ICS", "Derived", "Age Range", "", "N/A", "str", "utils.py", "Binned from Account Age Days"],
    # ── ICS Append: Output columns ──
    ["ICS Append", "Output", "ICS Account", "", "N/A", "str", "matcher.py", "Yes / No (appended to ODD)"],
    ["ICS Append", "Output", "ICS Source", "", "N/A", "str", "matcher.py", "REF / DM / Both / empty"],
    ["ICS Append", "Internal", "Acct Hash", "", "N/A", "str", "merger.py", "Normalized account hash from REF/DM files"],
    # ── ICS Referral: Required columns ──
    ["ICS Referral", "Required", "Referrer Name", "referrer, referrer_name", "Yes", "str", "Excel", "Person who made the referral"],
    ["ICS Referral", "Required", "Issue Date", "issue_date, date", "Yes", "datetime", "Excel", "Referral issue date"],
    ["ICS Referral", "Required", "Referral Code", "referral_code, code", "Yes", "str", "Excel", "Referral code identifier"],
    ["ICS Referral", "Required", "Purchase Manager", "purchase_manager, staff", "Yes", "str", "Excel", "Staff member / purchase manager"],
    ["ICS Referral", "Required", "Branch", "branch_id", "Yes", "str", "Excel", "Branch identifier"],
    ["ICS Referral", "Required", "Account Holder", "account_holder, new_account", "Yes", "str", "Excel", "New account holder name"],
    ["ICS Referral", "Required", "MRDB Account Hash", "mrdb_account_hash, mrdb, account_hash", "Yes", "str", "Excel", "Account hash for matching"],
    ["ICS Referral", "Required", "Cert ID", "cert_id, certificate_id", "Yes", "str", "Excel", "Certificate identifier"],
]

CROSS_PIPELINE_COLUMNS = [
    "ODD Field", "Used by ARS", "Used by TXN", "Used by ICS",
    "ARS Purpose", "TXN Purpose", "ICS Purpose",
]

CROSS_PIPELINE_DATA = [
    ["Acct Number", "No", "Yes (merge key)", "Yes (merge key)",
     "", "Joins TXN rows to ODD demographics", "Joins ICS append to ODD rows"],
    ["Stat Code", "Yes", "No", "Yes",
     "Subset filtering, eligibility", "", "Account status filtering"],
    ["Product Code / Prod Code", "Yes", "No", "Yes",
     "Eligibility, product analysis", "", "Product-level ICS analysis"],
    ["Date Opened", "Yes", "Yes (merge)", "Yes",
     "L12M filtering, age calc, attrition", "Account tenure", "Age calc, account age"],
    ["Date Closed", "Yes", "Yes (merge)", "Yes (optional)",
     "Open/closed split, attrition", "Account status", "Open/closed filtering"],
    ["Avg Bal", "Yes", "Yes (merge)", "Yes (optional)",
     "Balance categorization", "Balance tier binning", "Balance analysis"],
    ["Business?", "Yes", "Yes (merge)", "Yes",
     "Personal/business split", "Business flag on TXN rows", "P/B segmentation"],
    ["Debit?", "Yes", "Yes (merge)", "Yes",
     "DCTR, Reg E, Value modules", "Debit flag on TXN rows", "Debit account filtering"],
    ["Branch", "Yes", "Yes (merge)", "Yes",
     "Branch-level DCTR, attrition", "Branch on TXN rows", "Branch-level ICS analysis"],
    ["Account Holder Age", "Yes", "Yes (merge)", "No",
     "Demographics, age categories", "Generation cohort derivation", ""],
    ["DOB", "Yes", "No", "No",
     "Age computation in format_odd", "", ""],
    ["MmmYY Spend", "Yes", "No", "Yes",
     "Mailer impact, trends", "", "L12M activity analysis"],
    ["MmmYY Swipes", "Yes", "No", "Yes",
     "Mailer impact, trends", "", "L12M activity analysis"],
    ["MmmYY Mail", "Yes", "No", "No",
     "Mailer module, offer counting", "", ""],
    ["MmmYY Resp", "Yes", "Yes (segments)", "No",
     "Mailer module, segmentation", "Responder segment filters", ""],
    ["MmmYY Reg E Code", "Yes", "No", "No",
     "Reg E module analysis", "", ""],
    ["ICS Account", "No", "Yes (segments)", "Yes",
     "", "ICS segment filters", "Core analysis column"],
    ["ICS Source / Source", "No", "No", "Yes",
     "", "", "REF/DM/Both source analysis"],
    ["Curr Bal", "No", "No", "Yes",
     "", "", "Balance-tier analysis"],
]


def _write_generic_sheet(wb: Workbook, title: str, columns: list[str], data: list[list],
                         widths: list[int] | None = None) -> None:
    ws = wb.create_sheet(title=title)

    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=10)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=col_idx >= 3)

    if widths:
        for col_idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w
    else:
        for col_idx in range(1, len(columns) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(columns[col_idx - 1])
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 70)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(data) + 1}"


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    _write_sheet(wb, "ARS", ARS_DATA)
    _write_sheet(wb, "TXN", TXN_DATA)
    _write_sheet(wb, "ICS", ICS_DATA)
    _write_slide_type_sheet(wb)
    _write_layout_sheet(wb)
    _write_generic_sheet(wb, "Data Sources", DATA_SOURCE_COLUMNS, DATA_SOURCE_DATA,
                         [10, 18, 55, 18, 70, 70])
    _write_generic_sheet(wb, "M Drive Paths", M_DRIVE_COLUMNS, M_DRIVE_DATA,
                         [55, 40, 12, 30])
    _write_generic_sheet(wb, "File Paths", PATH_COLUMNS, PATH_DATA,
                         [16, 14, 80, 80])
    _write_generic_sheet(wb, "Field Names", FIELD_COLUMNS, FIELD_DATA,
                         [12, 14, 24, 55, 8, 10, 14, 60])
    _write_generic_sheet(wb, "Cross-Pipeline Fields", CROSS_PIPELINE_COLUMNS, CROSS_PIPELINE_DATA,
                         [24, 16, 16, 16, 35, 35, 35])

    out = "docs/slide_catalog.xlsx"
    wb.save(out)
    print(
        f"Saved {out} -- ARS: {len(ARS_DATA)} rows, TXN: {len(TXN_DATA)} rows, "
        f"ICS: {len(ICS_DATA)} rows, Layouts: {len(LAYOUT_DATA)} rows, "
        f"Paths: {len(PATH_DATA)} rows, Sources: {len(DATA_SOURCE_DATA)} rows, "
        f"M Drive: {len(M_DRIVE_DATA)} rows, "
        f"Fields: {len(FIELD_DATA)} rows, Cross-Pipeline: {len(CROSS_PIPELINE_DATA)} rows"
    )


if __name__ == "__main__":
    main()
