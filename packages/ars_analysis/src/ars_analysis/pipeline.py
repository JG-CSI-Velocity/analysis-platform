"""
pipeline.py
-----------
ARS Automated Reporting System — Pipeline Module

Refactored from v18-pre-analysis notebook (Cells 1-16 + Analyses A1-A5).
All variable names, function names, and terms are preserved exactly
as they appear in the original notebook.

Output locations:
    Source folder (beside ODD file):
        {client_id}-{year}-{month}-ars-analysis.xlsx
        {client_id}-{year}-{month}-{client_name}-presentation.pptx
        charts/  (chart images)
    Master Excel (cumulative):
        Presentations/Presentation Excels/{client_id}-master-{client_name}-ars-analysis.xlsx
    Archive (monthly copy):
        Presentations/Presentation Excels/Archive/{client_name}/{year}/{month}/

Usage:
    from pipeline import run_pipeline
    ctx = run_pipeline(file_path, config_path)
"""

# ============================
# Cell 1 — Imports
# ============================
import calendar
import json
import os
import re
import shutil
import sys
import time
import uuid
from datetime import datetime
from pathlib import Path

import matplotlib
import numpy as np
import pandas as pd

from ars_analysis.ars_config import ARCHIVE_PATH, PRESENTATIONS_PATH, TEMPLATE_PATH

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# PIPELINE CONTEXT
# =============================================================================


def create_context():
    """Create an empty pipeline context (mirrors notebook globals)."""
    return {
        "data": None,
        "data_original": None,
        "file_path": None,
        # Parsed from filename
        "client_id": None,
        "year": None,
        "month": None,
        "client_name": None,
        # Output paths
        "source_folder": None,
        "global_report_excel_file_path": None,
        "monthly_archive_file_path": None,
        "source_excel_path": None,  # ← Excel beside ODD file
        "pptx_output_path": None,  # ← PPTX beside ODD file
        "pptx_archive_path": None,
        # Config
        "client_config": None,
        "config": None,
        "eligible_stat_code": None,
        "eligible_prod_code": None,
        "eligible_mailable": None,
        "reg_e_opt_in": None,
        "nsf_od_fee": 0.0,
        "ic_rate": 0.0,
        "dctr_targets": {"peer_avg": 0.65, "p75": 0.72, "best_class": 0.80},
        "reg_e_target": 0.60,
        # Cleaned data
        "latest_reg_e_column": None,
        # Date ranges
        "start_date": None,
        "end_date": None,
        "last_12_months": None,
        # Subsets
        "open_accounts": None,
        "eligible_data": None,
        "eligible_personal": None,
        "eligible_business": None,
        "eligible_with_debit": None,
        "eligible_without_debit": None,
        "eligible_personal_with_debit": None,
        "eligible_business_with_debit": None,
        "open_personal": None,
        "open_business": None,
        "eligible_last_12m": None,
        "eligible_personal_last_12m": None,
        "eligible_business_last_12m": None,
        "open_last_12m": None,
        "reg_e_eligible_base": None,
        "reg_e_eligible_base_l12m": None,
        "reg_e_opted_in": None,
        "reg_e_opted_out": None,
        "closed_accounts": None,
        # Export
        "export_log": [],
        "chart_dir": None,
        # Deck builder
        "all_slides": [],
        "pptx_built": False,
        # Analysis results
        "results": {},
        # Progress callback (Streamlit)
        "_progress_callback": None,
        # Internal deck builder refs (set in step_setup_deck)
        "_deck_available": False,
        "_make_figure": None,
        "_SlideContent": None,
        "_DeckBuilder": None,
        "_DECK_CONFIG": None,
    }


def _report(ctx, message):
    """Print + optional Streamlit callback."""
    print(message)
    cb = ctx.get("_progress_callback")
    if cb:
        cb(message)


def _phase_time(ctx, label):
    """Log elapsed time since last phase checkpoint."""
    now = time.time()
    last = ctx.get("_phase_start", now)
    ctx["_phase_start"] = now
    if last != now:
        _report(ctx, f"   [{label}: {now - last:.1f}s]")


def _exec_report(ctx, method_name):
    """Call a set_* method on the executive report if present."""
    rpt = ctx.get("_exec_report")
    if rpt:
        try:
            getattr(rpt, method_name)(ctx)
        except Exception:
            pass  # report building must never crash the pipeline


def _exec_report_fail(ctx, section_key):
    """Mark a section as failed on the executive report."""
    rpt = ctx.get("_exec_report")
    if rpt:
        try:
            rpt.set_failed(section_key)
        except Exception:
            pass


# =============================================================================
# CELL 2 — Load ODD file
# =============================================================================


def step_load_data(ctx, file_path):
    """Cell 2: Load ODD file and create backup."""
    _report(ctx, f"📂 Loading {os.path.basename(file_path)}...")
    ctx["file_path"] = file_path
    ctx["source_folder"] = str(Path(file_path).parent)
    ctx["data"] = pd.read_excel(file_path)
    ctx["data_original"] = ctx["data"].copy()
    _report(ctx, f"📂 Loaded {len(ctx['data']):,} accounts with {len(ctx['data'].columns)} fields")
    return ctx


# =============================================================================
# CELL 3 — Parse filename + build paths
# =============================================================================


def parse_input_filename(file_path: str):
    """Parse 'ClientID-year-month-client name-ODD.xlsx' → tuple."""
    basename = os.path.basename(file_path)
    stem = os.path.splitext(basename)[0]
    parts = stem.split("-")
    if len(parts) < 5:
        raise ValueError(f"Expected 'ClientID-year-month-ClientName-ODD.xlsx', got: {basename}")
    if parts[-1].upper() != "ODD":
        raise ValueError(f"File must end with '-ODD.xlsx', got: {basename}")
    return parts[0], parts[1], parts[2], "-".join(parts[3:-1]).strip(), parts[-1]


def sanitize_filename_component(name: str) -> str:
    return re.sub(r'[<>:"/\\|?*]', "-", name).strip()


def sanitize_sheet_title(title: str) -> str:
    return re.sub(r"[:\\/\*?\[\]]", "-", title)[:31]


def atomic_create_empty_workbook(dest_path: str):
    p = Path(dest_path)
    if p.exists():
        return
    tmp = p.with_suffix(f".tmp.{uuid.uuid4().hex[:8]}.xlsx")
    try:
        p.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        wb.save(str(tmp))
        tmp.replace(p)
    except Exception:
        if tmp.exists():
            tmp.unlink()
        raise


def step_parse_and_paths(ctx, base_paths=None, metadata=None):
    """Cell 3: Parse filename, build ALL output paths."""
    file_path = ctx["file_path"]
    if metadata:
        client_id = metadata["client_id"]
        year = metadata["year"]
        month = metadata["month"]
        client_name = metadata.get("client_name", "")
    else:
        client_id, year, month, client_name, _ = parse_input_filename(file_path)
    ctx["client_id"] = client_id
    ctx["year"] = year
    ctx["month"] = month
    ctx["client_name"] = client_name

    client_name_fs = sanitize_filename_component(client_name)
    source_folder = Path(file_path).parent

    if base_paths is None:
        base_paths = {
            "presentations": PRESENTATIONS_PATH,
            "archive": ARCHIVE_PATH,
        }

    # (A) Master workbook (cumulative across months)
    master_file_name = f"{client_id}-master-{client_name_fs}-ars-analysis.xlsx"
    global_report_excel_file_path = str(base_paths["presentations"] / master_file_name)

    # (B) Monthly archive workbook
    monthly_archive_file_name = f"{client_id}-{year}-{month}-ars-analysis.xlsx"
    month_folder = base_paths["archive"] / client_name_fs / year / month
    month_folder.mkdir(parents=True, exist_ok=True)
    monthly_archive_file_path = str(month_folder / monthly_archive_file_name)

    # (C) Source folder outputs (beside the ODD file — PRIMARY outputs)
    source_excel_path = str(source_folder / monthly_archive_file_name)
    pptx_output_path = str(
        source_folder / f"{client_id}-{year}-{month}-{client_name}-presentation.pptx"
    )
    pptx_archive_path = str(
        month_folder / f"{client_id}-{year}-{month}-{client_name}-presentation.pptx"
    )

    # Chart directory
    chart_dir = source_folder / "charts"
    chart_dir.mkdir(parents=True, exist_ok=True)

    # Create empty workbooks
    atomic_create_empty_workbook(global_report_excel_file_path)
    atomic_create_empty_workbook(source_excel_path)

    ctx.update(
        {
            "global_report_excel_file_path": global_report_excel_file_path,
            "monthly_archive_file_path": monthly_archive_file_path,
            "source_excel_path": source_excel_path,
            "pptx_output_path": pptx_output_path,
            "pptx_archive_path": pptx_archive_path,
            "chart_dir": chart_dir,
        }
    )

    _report(ctx, f"📁 Source folder: {source_folder}")
    _report(ctx, f"📊 Excel output:  {source_excel_path}")
    _report(ctx, f"📽️ PPTX output:   {pptx_output_path}")
    _report(ctx, f"📁 Master Excel:  {global_report_excel_file_path}")
    return ctx


# =============================================================================
# CELLS 4-6 — Load config
# =============================================================================


def step_load_config(ctx, config_path=None):
    """Cells 4-6: Load JSON config, get client settings."""
    if config_path is None:
        from ars_analysis.ars_config import CONFIG_PATH

        config_path = str(CONFIG_PATH)

    with open(config_path) as f:
        client_config = json.load(f)
    ctx["client_config"] = client_config

    client_id = ctx["client_id"]
    config = client_config.get(client_id)
    if config is None:
        available = list(client_config.keys())[:10]
        raise ValueError(f"No config for client '{client_id}'. Available: {', '.join(available)}")
    ctx["config"] = config

    # Extract config values (ensure lists)
    def ensure_list(val):
        return [val] if isinstance(val, str) else val

    ctx["eligible_stat_code"] = ensure_list(config["EligibleStatusCodes"])
    ctx["eligible_prod_code"] = ensure_list(config["EligibleProductCodes"])
    ctx["eligible_mailable"] = ensure_list(config["EligibleMailCode"])
    ctx["reg_e_opt_in"] = ensure_list(config["RegEOptInCode"])

    # Numeric values
    try:
        ctx["ic_rate"] = float(str(config.get("ICRate", "0")).strip())
    except (ValueError, TypeError):
        ctx["ic_rate"] = 0.0
    try:
        ctx["nsf_od_fee"] = float(str(config.get("NSF_OD_Fee", "0")).strip())
    except (ValueError, TypeError):
        ctx["nsf_od_fee"] = 0.0

    # Benchmark targets (optional, per-client)
    _DEFAULT_DCTR_TARGETS = {"peer_avg": 0.65, "p75": 0.72, "best_class": 0.80}
    benchmarks = config.get("benchmarks", {})
    dctr_targets = benchmarks.get("dctr_targets", _DEFAULT_DCTR_TARGETS)
    # Validate: all values must be fractions in (0, 1]
    validated_targets = {}
    for key in ("peer_avg", "p75", "best_class"):
        val = dctr_targets.get(key, _DEFAULT_DCTR_TARGETS[key])
        try:
            val = float(val)
        except (ValueError, TypeError):
            val = _DEFAULT_DCTR_TARGETS[key]
        if not (0 < val <= 1):
            _report(
                ctx, f"   ⚠️ benchmarks.dctr_targets.{key}={val} out of range (0,1], using default"
            )
            val = _DEFAULT_DCTR_TARGETS[key]
        validated_targets[key] = val
    ctx["dctr_targets"] = validated_targets
    ctx["reg_e_target"] = float(benchmarks.get("reg_e_target", 0.60))

    _report(ctx, f"✅ Config loaded for {client_id} — {ctx['client_name']}")
    _report(ctx, f"   NSF/OD Fee: ${ctx['nsf_od_fee']:.2f}  |  IC Rate: {ctx['ic_rate']:.4%}")
    _report(
        ctx,
        f"   Stat codes: {len(ctx['eligible_stat_code'])}  |  Prod codes: {len(ctx['eligible_prod_code'])}",
    )
    if benchmarks:
        _report(ctx, f"   DCTR targets: {validated_targets}")
    return ctx


# =============================================================================
# CELL 7 — Clean data
# =============================================================================


def clean_code_column(val):
    try:
        f = float(val)
        i = int(f)
        return str(i) if f == i else str(f)
    except (ValueError, TypeError):
        return str(val).strip() if pd.notna(val) else ""


def safe_clean_column(df, column_name, cleaning_func=None):
    if column_name not in df.columns:
        return df
    try:
        if cleaning_func:
            df[column_name] = df[column_name].apply(cleaning_func)
        else:
            df[column_name] = df[column_name].astype(str).str.strip()
        df[column_name] = df[column_name].replace("nan", "")
        return df
    except Exception as e:
        print(f"❌ Error cleaning '{column_name}': {e}")
        return df


def step_clean_data(ctx):
    """Cell 7: Clean columns, identify Reg E, quality report."""
    data = ctx["data"]
    _report(ctx, f"🧼 Cleaning {len(data):,} rows...")

    data = safe_clean_column(data, "Prod Code", clean_code_column)
    data = safe_clean_column(data, "Branch", clean_code_column)
    data = safe_clean_column(data, "Stat Code")
    data = safe_clean_column(data, "Mailable?")
    data = safe_clean_column(data, "Debit?")

    # Quality flags
    missing_prod = (data["Prod Code"] == "") | (data["Prod Code"].str.lower() == "nan")
    missing_branch = (data["Branch"] == "") | (data["Branch"].str.lower() == "nan")
    data["Has_Valid_ProdCode"] = ~missing_prod
    data["Has_Valid_Branch"] = ~missing_branch

    # Reg E column
    try:
        reg_e_cols = [c for c in data.columns if "Reg E Code" in c]
        if reg_e_cols:
            latest = sorted(
                reg_e_cols,
                key=lambda x: pd.to_datetime(x.split(" ")[0], format="%b%y", errors="coerce"),
                reverse=True,
            )[0]
            data = safe_clean_column(data, latest)
            ctx["latest_reg_e_column"] = latest
            _report(ctx, f"   Reg E column: {latest}")
        else:
            ctx["latest_reg_e_column"] = None
            _report(ctx, "   ⚠️ No Reg E columns found")
    except Exception:
        ctx["latest_reg_e_column"] = None

    ctx["data"] = data
    _report(ctx, f"✅ Cleaning complete — {len(data):,} rows retained")
    return ctx


# =============================================================================
# CELL 8 — Date range
# =============================================================================


def step_date_range(ctx):
    """Cell 8: Calculate L12M period."""
    now = datetime.now()
    end_date = pd.Timestamp(now.replace(day=1)) - pd.DateOffset(days=1)
    start_date = end_date.replace(day=1) - pd.DateOffset(months=11)
    last_12_months = (
        pd.date_range(start=start_date, end=end_date, freq="ME").strftime("%b%y").tolist()
    )

    ctx["start_date"] = start_date
    ctx["end_date"] = end_date
    ctx["last_12_months"] = last_12_months

    _report(
        ctx,
        f"📅 Analysis Period: {start_date.strftime('%B %Y')} – {end_date.strftime('%B %Y')} ({len(last_12_months)} months)",
    )
    return ctx


# =============================================================================
# CELL 11 — ARSFieldFormatter
# =============================================================================


class ARSFieldFormatter:
    FORMATTING_RULES = {
        "currency": {
            "patterns": ["$", "spend", "bal", "limit", "fee", "amount", "revenue"],
            "excel_format": "$#,##0.00",
        },
        "integer": {
            "patterns": ["swipes", "#", "count", "items", "mtd", "transactions"],
            "excel_format": "#,##0",
        },
        "percentage": {
            "patterns": ["rate", "%", "percent", "ratio", "pct", "dctr"],
            "excel_format": "0.00%",
        },
        "decimal": {
            "patterns": ["age", "years", "months", "days", "score"],
            "excel_format": "0.00",
        },
        "text": {
            "patterns": ["id", "code", "desc", "mask", "name", "branch", "type"],
            "excel_format": "@",
        },
        "date": {"patterns": ["date", "dob", "opened", "closed"], "excel_format": "mm/dd/yyyy"},
        "status": {"patterns": ["?", "elig", "mailable", "debit", "business"], "excel_format": "@"},
    }

    @classmethod
    def identify_column_format(cls, col_name):
        col_lower = col_name.lower().strip()
        exact = {"branch": "text", "age": "decimal", "rate": "percentage"}
        if col_lower in exact:
            return exact[col_lower]
        for fmt, rules in cls.FORMATTING_RULES.items():
            for p in rules["patterns"]:
                if p in col_lower:
                    return fmt
        return "text"


# =============================================================================
# CELLS 12 — Excel export system
# =============================================================================


def _get_workbook(ctx, file_path):
    """Return a cached workbook for file_path, loading it once."""
    cache = ctx.setdefault("_open_workbooks", {})
    key = str(file_path)
    if key not in cache:
        try:
            cache[key] = load_workbook(key)
        except Exception:
            cache[key] = Workbook()
    return cache[key]


def flush_workbooks(ctx):
    """Save all cached workbooks to disk (called once at end of pipeline)."""
    for path, wb in ctx.get("_open_workbooks", {}).items():
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]
        safe_save_workbook(wb, path)
    ctx["_open_workbooks"] = {}


def safe_save_workbook(wb, file_path, max_retries=8, sleep_s=0.5):
    target = Path(file_path)
    tmp = target.with_suffix(f".tmp.{uuid.uuid4().hex[:8]}.xlsx")
    for attempt in range(1, max_retries + 1):
        try:
            wb.save(str(tmp))
            tmp.replace(target)
            return
        except PermissionError:
            if attempt < max_retries:
                time.sleep(sleep_s * attempt)
            else:
                raise
        except Exception:
            if tmp.exists():
                tmp.unlink()
            raise


def write_formatted_df(ws, df, row, wb, sheet_title=None, key_metrics=None):
    """Write DataFrame to worksheet with formatting. Returns last data row."""
    if sheet_title:
        cell = ws.cell(row=row, column=1, value=sheet_title)
        cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if df is not None and isinstance(df, pd.DataFrame) and len(df.columns) > 0:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(df.columns))
        row += 2

    if key_metrics and isinstance(key_metrics, dict):
        for k, v in key_metrics.items():
            ws.cell(row=row, column=1, value=k).font = Font(bold=True)
            ws.cell(row=row, column=2, value=str(v))
            row += 1
        row += 1

    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return row

    # Headers
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=row, column=ci, value=col)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin")
        )

    # Data
    for ri, row_data in enumerate(df.itertuples(index=False), row + 1):
        for ci, (col, val) in enumerate(zip(df.columns, row_data), 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            fmt = ARSFieldFormatter.identify_column_format(col)
            if fmt in ARSFieldFormatter.FORMATTING_RULES:
                cell.number_format = ARSFieldFormatter.FORMATTING_RULES[fmt]["excel_format"]
            cell.border = Border(
                left=Side("thin", color="D9D9D9"),
                right=Side("thin", color="D9D9D9"),
                top=Side("thin", color="D9D9D9"),
                bottom=Side("thin", color="D9D9D9"),
            )

    end_row = row + len(df)

    # Autosize
    for ci in range(1, len(df.columns) + 1):
        letter = get_column_letter(ci)
        max_len = max(
            len(str(df.columns[ci - 1])),
            max(
                (
                    len(str(ws.cell(row=r, column=ci).value or ""))
                    for r in range(row + 1, end_row + 1)
                ),
                default=0,
            ),
        )
        ws.column_dimensions[letter].width = min(max_len + 2, 50)

    return end_row


def _init_summary_sheet(ws):
    ws.cell(row=1, column=1, value="ARS Analysis Report Summary")
    c = ws.cell(row=1, column=1)
    c.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    c.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells("A1:F1")
    for ci, h in enumerate(["#", "Analysis", "Sheet", "Time", "Rows", "Metrics"], 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")


def save_to_excel(ctx, df, sheet_name, analysis_title, key_metrics=None):
    """Write to source Excel + master Excel (deferred save via flush_workbooks)."""
    if ctx.get("_skip_excel"):
        return
    sheet_title = sanitize_sheet_title(sheet_name)

    for file_path in [ctx["source_excel_path"], ctx["global_report_excel_file_path"]]:
        wb = _get_workbook(ctx, file_path)

        if sheet_title in wb.sheetnames:
            del wb[sheet_title]

        ws = wb.create_sheet(title=sheet_title)

        if isinstance(df, dict):
            cur_row = 1
            for sub_name, sub_df in df.items():
                if sub_df is not None and isinstance(sub_df, pd.DataFrame) and not sub_df.empty:
                    end = write_formatted_df(
                        ws,
                        sub_df,
                        cur_row,
                        wb,
                        sheet_title=f"{analysis_title} — {sub_name}",
                        key_metrics=key_metrics if cur_row == 1 else None,
                    )
                    cur_row = end + 3
        else:
            write_formatted_df(ws, df, 1, wb, sheet_title=analysis_title, key_metrics=key_metrics)

        # Summary sheet
        if "Summary" not in wb.sheetnames:
            sws = wb.create_sheet(title="Summary", index=0)
            _init_summary_sheet(sws)
        else:
            sws = wb["Summary"]
        nr = sws.max_row + 1
        rc = (
            len(df)
            if isinstance(df, pd.DataFrame)
            else sum(len(v) for v in df.values() if isinstance(v, pd.DataFrame))
        )
        ms = " | ".join(f"{k}: {v}" for k, v in key_metrics.items()) if key_metrics else ""
        sws.cell(row=nr, column=1, value=nr - 2)
        sws.cell(row=nr, column=2, value=analysis_title)
        sws.cell(row=nr, column=3, value=sheet_title)
        sws.cell(row=nr, column=4, value=datetime.now().strftime("%H:%M"))
        sws.cell(row=nr, column=5, value=rc)
        sws.cell(row=nr, column=6, value=ms)

    ctx["export_log"].append({"analysis": sheet_name, "time": datetime.now().strftime("%H:%M:%S")})


# =============================================================================
# CELL 14 — Create common data subsets
# =============================================================================


def step_create_subsets(ctx):
    """Cell 14: Create all common data subsets."""
    data = ctx["data"]
    _report(ctx, "\n📊 CREATING DATA SUBSETS")

    esc = ctx["eligible_stat_code"]
    epc = ctx["eligible_prod_code"]
    em = (
        ctx["eligible_mailable"]
        if isinstance(ctx["eligible_mailable"], list)
        else [ctx["eligible_mailable"]]
    )
    sd, ed = ctx["start_date"], ctx["end_date"]

    # Open and closed accounts
    open_accounts = data[data["Date Closed"].isna()].copy()
    closed_accounts = data[data["Date Closed"].notna()].copy()
    _report(ctx, f"   Open accounts: {len(open_accounts):,}")
    _report(ctx, f"   Closed accounts: {len(closed_accounts):,}")

    # Eligible
    eligible_data = data[
        (data["Date Closed"].isna())
        & (data["Stat Code"].isin(esc))
        & (data["Prod Code"].isin(epc))
        & (data["Mailable?"].isin(em))
        & (data["Avg Bal"] != 0)
    ].copy()
    _report(ctx, f"   Eligible accounts: {len(eligible_data):,}")

    # Business type
    eligible_personal = eligible_data[eligible_data["Business?"] == "No"].copy()
    eligible_business = eligible_data[eligible_data["Business?"] == "Yes"].copy()
    open_personal = open_accounts[open_accounts["Business?"] == "No"].copy()
    open_business = open_accounts[open_accounts["Business?"] == "Yes"].copy()

    # Debit card
    eligible_with_debit = eligible_data[eligible_data["Debit?"] == "Yes"].copy()
    eligible_without_debit = eligible_data[eligible_data["Debit?"] == "No"].copy()
    eligible_personal_with_debit = eligible_personal[eligible_personal["Debit?"] == "Yes"].copy()
    eligible_business_with_debit = eligible_business[eligible_business["Debit?"] == "Yes"].copy()

    _report(
        ctx,
        f"   With debit: {len(eligible_with_debit):,}  |  Without: {len(eligible_without_debit):,}",
    )
    _report(
        ctx, f"   Personal: {len(eligible_personal):,}  |  Business: {len(eligible_business):,}"
    )

    # L12M subsets
    for df_ref in [eligible_data, eligible_personal, eligible_business, open_accounts]:
        if "Date Opened" in df_ref.columns:
            df_ref["Date Opened"] = pd.to_datetime(df_ref["Date Opened"], errors="coerce")

    def l12m_filter(df):
        return df[(df["Date Opened"] >= sd) & (df["Date Opened"] <= ed)].copy()

    eligible_last_12m = l12m_filter(eligible_data)
    eligible_personal_last_12m = l12m_filter(eligible_personal)
    eligible_business_last_12m = l12m_filter(eligible_business)
    open_last_12m = l12m_filter(open_accounts)

    _report(ctx, f"   L12M eligible: {len(eligible_last_12m):,}")

    # Reg E
    reg_e_eligible_base = reg_e_eligible_base_l12m = reg_e_opted_in = reg_e_opted_out = None
    lrc = ctx["latest_reg_e_column"]
    if lrc:
        reg_e_eligible_base = eligible_personal[eligible_personal["Debit?"] == "Yes"].copy()
        reg_e_eligible_base[lrc] = reg_e_eligible_base[lrc].astype(str).str.strip()
        opt_list = [str(v).strip() for v in ctx["reg_e_opt_in"]]
        reg_e_opted_in = reg_e_eligible_base[reg_e_eligible_base[lrc].isin(opt_list)].copy()
        reg_e_opted_out = reg_e_eligible_base[~reg_e_eligible_base[lrc].isin(opt_list)].copy()
        reg_e_eligible_base_l12m = eligible_personal_last_12m[
            eligible_personal_last_12m["Debit?"] == "Yes"
        ].copy()
        _report(
            ctx, f"   Reg E: {len(reg_e_opted_in):,} opted in / {len(reg_e_opted_out):,} opted out"
        )

    ctx.update(
        {
            "open_accounts": open_accounts,
            "closed_accounts": closed_accounts,
            "eligible_data": eligible_data,
            "eligible_personal": eligible_personal,
            "eligible_business": eligible_business,
            "eligible_with_debit": eligible_with_debit,
            "eligible_without_debit": eligible_without_debit,
            "eligible_personal_with_debit": eligible_personal_with_debit,
            "eligible_business_with_debit": eligible_business_with_debit,
            "open_personal": open_personal,
            "open_business": open_business,
            "eligible_last_12m": eligible_last_12m,
            "eligible_personal_last_12m": eligible_personal_last_12m,
            "eligible_business_last_12m": eligible_business_last_12m,
            "open_last_12m": open_last_12m,
            "reg_e_eligible_base": reg_e_eligible_base,
            "reg_e_eligible_base_l12m": reg_e_eligible_base_l12m,
            "reg_e_opted_in": reg_e_opted_in,
            "reg_e_opted_out": reg_e_opted_out,
        }
    )
    return ctx


# =============================================================================
# CELL 16 — Deck builder setup
# =============================================================================


def step_setup_deck(ctx, deck_builder_path=None, template_path=None):
    """Cell 16: Initialize deck builder."""

    try:
        from ars_analysis.deck_builder import (
            DECK_CONFIG,
            DeckBuilder,
            SlideContent,
            apply_matplotlib_defaults,
            make_figure,
            setup_slide_helpers,  # noqa: F401
        )

        ctx["_deck_available"] = True
        ctx["_SlideContent"] = SlideContent
        ctx["_DeckBuilder"] = DeckBuilder
        ctx["_make_figure"] = make_figure
        ctx["_DECK_CONFIG"] = DECK_CONFIG

        if template_path is None:
            template_path = str(TEMPLATE_PATH)
        DECK_CONFIG["template_path"] = template_path

        apply_matplotlib_defaults()
        _report(ctx, "✅ Deck builder loaded")
    except ImportError as e:
        _report(ctx, f"⚠️ Deck builder not available: {e}")
        ctx["_deck_available"] = False
        ctx["_make_figure"] = None
    return ctx


def _get_figure(ctx, size="single"):
    """Get a matplotlib figure — uses deck_builder's make_figure if available."""
    mf = ctx.get("_make_figure")
    if mf:
        return mf(size)
    sizes = {"single": (10, 5), "half": (6, 4), "wide": (14, 6), "square": (7, 7)}
    return plt.subplots(figsize=sizes.get(size, (10, 5)))


def add_to_library(ctx, slide_id, slide_data, category="General"):
    """Add a slide to the library."""
    ctx["all_slides"].append(
        {"id": slide_id, "category": category, "data": slide_data, "include": True}
    )


def build_slide_content(ctx, slide_data, default_layout=4):
    """Convert insight dict → SlideContent."""
    SC = ctx.get("_SlideContent")
    if SC is None:
        return None
    if isinstance(slide_data, SC):
        return slide_data

    title = slide_data.get("title", "Analysis")
    subtitle = slide_data.get("subtitle", "")
    layout_index = slide_data.get("layout_index", default_layout)
    full_title = f"{title}\n{subtitle}" if subtitle else title

    # --- Mailer summary composite slide ---
    if slide_data.get("slide_type") == "mailer_summary":
        donut = slide_data.get("donut_path", "")
        hbar = slide_data.get("hbar_path", "")
        kpis = slide_data.get("kpis", {})
        inside_numbers = slide_data.get("inside_numbers", [])
        insight_text = slide_data.get("insight_text", "")

        # Pack inside_numbers as "pct|description" strings in bullets
        bullets = [insight_text]
        for item in inside_numbers:
            if isinstance(item, (list, tuple)) and len(item) == 2:
                bullets.append(f"{item[0]}|{item[1]}")
            else:
                bullets.append(str(item))

        return SC(
            slide_type="mailer_summary",
            title=title,
            images=[donut, hbar],
            kpis=kpis,
            bullets=bullets,
            layout_index=layout_index,
        )

    # --- Multi-screenshot (two charts side by side) ---
    if slide_data.get("slide_type") == "multi_screenshot":
        imgs = [slide_data.get("chart_path", ""), slide_data.get("chart_path_2", "")]
        return SC(
            slide_type="multi_screenshot", title=full_title, images=imgs, layout_index=layout_index
        )

    # --- Standard chart slides ---
    kpis = slide_data.get("kpis", {})
    chart_path = slide_data.get("chart_path")

    if not chart_path:
        return None

    if kpis:
        return SC(
            slide_type="screenshot_kpi",
            title=full_title,
            images=[chart_path],
            kpis=kpis,
            layout_index=layout_index,
        )
    else:
        return SC(
            slide_type="screenshot",
            title=full_title,
            images=[chart_path],
            layout_index=layout_index,
        )


# =============================================================================
# PREAMBLE SLIDES — Non-analysis placeholder slides at start of deck
# =============================================================================


def _build_preamble_slides(ctx):
    """
    Build the 12 preamble slides that precede analysis content.

    These are a mix of:
      - Title/intro slides (layout 0, 1)
      - Section dividers (layout 1, 2, 3)
      - Blank placeholders for manual content (layout 3, 8, 15)
      - Analysis chart slots (layout 15) — left blank here,
        wired to modules later

    Returns list of SlideContent objects in presentation order.
    """
    SC = ctx["_SlideContent"]
    month_name = calendar.month_name[int(ctx["month"])]
    title_date = f"{month_name} {ctx['year']}"
    client_name = ctx["client_name"]

    preamble = [
        # 1. Intro — client name + month, white text on dark bg
        SC(
            slide_type="title",
            title=f"{client_name}\nAccount Revenue Solution | {title_date}",
            layout_index=1,
        ),
        # 2. Agenda — dedicated Agenda layout in network template
        SC(slide_type="blank", title="Agenda", layout_index=14),
        # 3. Program Performance — section divider (client name + section | month)
        SC(
            slide_type="title",
            title=f"{client_name}\nProgram Performance | {title_date}",
            layout_index=1,
        ),
        # 4. Financial Performance — blank for manual table
        SC(slide_type="blank", title="Financial Performance", layout_index=0),
        # 5. Monthly Revenue – Last 12 Months — blank for manual content
        SC(slide_type="blank", title="Monthly Revenue – Last 12 Months", layout_index=12),
        # 6. ARS Lift Matrix — blank placeholder
        SC(slide_type="blank", title="ARS Lift Matrix", layout_index=8),
        # 7. ARS Mailer Revisit — section sub-divider (R07)
        SC(
            slide_type="title",
            title=f"{client_name}\nARS Mailer Revisit | {title_date}",
            layout_index=1,
        ),
        # 8. ARS Mailer Revisit – Swipes — placeholder (will wire to A12 later)
        SC(slide_type="blank", title="ARS Mailer Revisit – Swipes", layout_index=13),
        # 9. ARS Mailer Revisit – Spend — placeholder (will wire to A12 later)
        SC(slide_type="blank", title="ARS Mailer Revisit – Spend", layout_index=13),
        # 10. DCO — blank
        SC(
            slide_type="blank",
            title="Data Check Overview\nOur goal is turning non-users and light-users into heavy users",
            layout_index=8,
        ),
        # 11. Mailer Summaries — section divider
        SC(
            slide_type="title",
            title=f"Mailer Summaries\n{client_name} | {title_date}",
            layout_index=2,
        ),
        # 12. All Program Results — blank
        SC(slide_type="blank", title="All Program Results", layout_index=2),
        # 13. Program Responses to Date (A13.5) — placeholder (will wire later)
        SC(slide_type="blank", title="Program Responses to Date", layout_index=13),
    ]

    _report(ctx, f"📋 Built {len(preamble)} preamble slides")
    return preamble


# =============================================================================
# DECK BUILD — Actually creates the PowerPoint
# =============================================================================


def _reorder_analysis_slides(ctx, selected):
    """
    Reorder analysis slides to match reference section flow.

    Target order:
      1. 2 most recent mailer clusters (Summary + Swipes + Spend each)
      2. A1-A5 Overview
      3. A7.x DCTR + A11.1 Value of a Debit Card
      4. A8.x Reg E + A11.2 Value of Reg E Opt-In
      5. Remaining mailer clusters (newest first) — appendix
      6. A13 All-Time Summary + A14.2
    """
    overview = []
    dctr = []
    reg_e = []
    value = []
    attrition = []
    mailer_a12 = {}
    mailer_a13_monthly = {}
    mailer_agg = []
    other = []

    for s in selected:
        sid = s["id"]
        cat = s["category"]

        if cat == "Overview":
            overview.append(s)
        elif cat == "DCTR":
            dctr.append(s)
        elif cat == "Reg E":
            reg_e.append(s)
        elif cat == "Value":
            value.append(s)
        elif cat == "Attrition":
            attrition.append(s)
        elif sid.startswith("A12 - "):
            # "A12 - Nov25 Swipes" → month="Nov25", type="Swipes"
            parts = sid.replace("A12 - ", "").rsplit(" ", 1)
            if len(parts) == 2:
                month, stype = parts
                mailer_a12.setdefault(month, {})[stype] = s
            else:
                other.append(s)
        elif sid.startswith("A13 - ") and "All-Time" not in sid:
            # "A13 - Nov25 Mailer Summary" → month="Nov25"
            m = re.match(r"A13 - (\w+) Mailer Summary", sid)
            if m:
                mailer_a13_monthly[m.group(1)] = s
            else:
                other.append(s)
        elif sid.startswith(("A13 - All-Time", "A13.", "A14.", "A15.")):
            mailer_agg.append(s)
        else:
            other.append(s)

    # Sort months chronologically
    all_months = sorted(
        set(list(mailer_a13_monthly.keys()) + list(mailer_a12.keys())),
        key=lambda m: pd.to_datetime(m, format="%b%y", errors="coerce"),
    )

    recent_2 = list(reversed(all_months[-2:])) if len(all_months) >= 2 else all_months
    appendix_months = all_months[:-2] if len(all_months) > 2 else []
    # Appendix in newest-first order (matches reference)
    appendix_months = list(reversed(appendix_months))

    def _cluster(month):
        """Build a [Summary, Swipes, Spend] cluster for a month."""
        cluster = []
        if month in mailer_a13_monthly:
            cluster.append(mailer_a13_monthly[month])
        a12 = mailer_a12.get(month, {})
        if "Swipes" in a12:
            cluster.append(a12["Swipes"])
        if "Spend" in a12:
            cluster.append(a12["Spend"])
        return cluster

    recent_clusters = []
    for month in recent_2:
        recent_clusters.extend(_cluster(month))

    appendix_clusters = []
    for month in appendix_months:
        appendix_clusters.extend(_cluster(month))

    # Append each value slide to its respective section
    value_dctr = [s for s in value if s["id"] == "A11.1 - Value of a Debit Card"]
    value_reg_e = [s for s in value if s["id"] == "A11.2 - Value of Reg E Opt-In"]

    # ------------------------------------------------------------------
    # Consolidation: merge paired slides and move detail to appendix
    # ------------------------------------------------------------------
    from ars_analysis.dctr import DCTR_APPENDIX_IDS, DCTR_MERGES
    from ars_analysis.reg_e import REGE_APPENDIX_IDS, REGE_MERGES

    def _consolidate(slides, merges, appendix_ids):
        """Merge paired slides and separate appendix slides.

        Walks the original slide list in narrative order:
        - Left side of a merge pair -> insert merged multi-screenshot slide
        - Right side of a merge pair -> skip (already included)
        - Appendix ID -> move to appendix
        - Everything else -> keep as single slide
        """
        # Build lookup: left_id -> merged dict, right_id -> skip
        merge_at = {}  # left_id -> merged slide dict
        skip_ids = set()
        by_id = {s["id"]: s for s in slides}

        for left_id, right_id, title in merges:
            left = by_id.get(left_id)
            right = by_id.get(right_id)
            if left and right:
                merge_at[left_id] = {
                    "id": f"{left_id} + {right_id}",
                    "category": left["category"],
                    "data": {
                        "title": title,
                        "chart_path": left["data"].get("chart_path", ""),
                        "chart_path_2": right["data"].get("chart_path", ""),
                        "slide_type": "multi_screenshot",
                        "layout_index": 6,
                    },
                }
                skip_ids.add(left_id)
                skip_ids.add(right_id)

        result = []
        appendix_out = []
        for s in slides:
            sid = s["id"]
            if sid in merge_at:
                result.append(merge_at[sid])
            elif sid in skip_ids:
                continue
            elif sid in appendix_ids:
                appendix_out.append(s)
            else:
                result.append(s)

        return result, appendix_out

    dctr_main, dctr_appendix = _consolidate(dctr, DCTR_MERGES, DCTR_APPENDIX_IDS)
    rege_main, rege_appendix = _consolidate(reg_e, REGE_MERGES, REGE_APPENDIX_IDS)

    ATTRITION_MERGES = [
        (
            "A9.3 - Open vs Closed",
            "A9.6 - Personal vs Business",
            "Attrition Profile: Open vs Closed & Personal vs Business",
        ),
    ]
    ATTRITION_APPENDIX_IDS = {
        "A9.2 - Closure Duration",
        "A9.4 - Attrition by Branch",
        "A9.5 - Attrition by Product",
        "A9.7 - Attrition by Tenure",
        "A9.8 - Attrition by Balance",
        "A9.13 - ARS vs Non-ARS",
    }
    attrition_main, attrition_appendix = _consolidate(
        attrition, ATTRITION_MERGES, ATTRITION_APPENDIX_IDS
    )

    _report(
        ctx,
        f"   Consolidated DCTR: {len(dctr)} -> {len(dctr_main)} main + {len(dctr_appendix)} appendix",
    )
    _report(
        ctx,
        f"   Consolidated Reg E: {len(reg_e)} -> {len(rege_main)} main + {len(rege_appendix)} appendix",
    )
    _report(
        ctx,
        f"   Consolidated Attrition: {len(attrition)} -> {len(attrition_main)} main + {len(attrition_appendix)} appendix",
    )

    # Build subtitle for section dividers: "Client Name | Mon YYYY"
    month_name = calendar.month_name[int(ctx["month"])]
    section_subtitle = f"{ctx['client_name']} | {month_name} {ctx['year']}"

    def _section(title, layout_index=1, slide_type="title", kpis=None, subtitle=None):
        """Create a section divider marker dict."""
        full_title = f"{title}\n{subtitle}" if subtitle else title
        return {
            "__section__": True,
            "id": f"__section__{title}",
            "slide_type": slide_type,
            "title": full_title,
            "layout_index": layout_index,
            "kpis": kpis or {},
        }

    ordered = list(recent_clusters)

    if dctr_main or value_dctr:
        ordered.append(_section("Debit Card Take Rate", subtitle=section_subtitle))
        ordered.extend(dctr_main)
        ordered.extend(value_dctr)

    if rege_main or value_reg_e:
        ordered.append(_section("Reg E Analysis", subtitle=section_subtitle))
        ordered.extend(rege_main)
        ordered.extend(value_reg_e)

    if attrition_main:
        ordered.append(_section("Account Attrition", subtitle=section_subtitle))
        ordered.extend(attrition_main)

    # Summary placeholder after analysis, before appendix
    ordered.append(_section("Summary & Key Takeaways", layout_index=12, slide_type="blank"))

    # Appendix: divider + older mailer months + overview + analysis detail
    has_appendix = (
        appendix_clusters or dctr_appendix or rege_appendix or attrition_appendix or overview
    )
    if has_appendix:
        ordered.append(_section("Appendix", subtitle=section_subtitle))
        ordered.extend(appendix_clusters)
        if overview:
            ordered.extend(overview)
        if dctr_appendix:
            ordered.extend(dctr_appendix)
        if rege_appendix:
            ordered.extend(rege_appendix)
        if attrition_appendix:
            ordered.extend(attrition_appendix)

    ordered.extend(mailer_agg)
    ordered.extend(other)

    _report(
        ctx,
        f"   Section order: "
        f"{len(recent_clusters)} recent mailer, "
        f"{len(overview)} overview, "
        f"{len(dctr_main)}+{len(value_dctr)} DCTR, "
        f"{len(rege_main)}+{len(value_reg_e)} Reg E, "
        f"{len(attrition_main)} attrition, "
        f"{len(appendix_clusters)} appendix mailer, "
        f"{len(dctr_appendix)}+{len(rege_appendix)}"
        f"+{len(attrition_appendix)} analysis appendix, "
        f"{len(mailer_agg)} aggregate, "
        f"{len(other)} other",
    )

    # Cross-category merge: L12M Funnel + most recent A12 Spend
    funnel_idx = None
    spend_idx = None
    for i, s in enumerate(ordered):
        sid = s.get("id", "")
        if sid == "A7.8 - L12M Funnel":
            funnel_idx = i
        elif sid.startswith("A12 - ") and sid.endswith(" Spend"):
            if spend_idx is None:
                spend_idx = i

    if funnel_idx is not None and spend_idx is not None:
        funnel_s = ordered[funnel_idx]
        spend_s = ordered[spend_idx]
        merged = {
            "id": f"{funnel_s['id']} + {spend_s['id']}",
            "category": funnel_s.get("category", "DCTR"),
            "data": {
                "title": "L12M Funnel & Mail Campaign Spend",
                "chart_path": funnel_s["data"].get("chart_path", ""),
                "chart_path_2": spend_s["data"].get("chart_path", ""),
                "slide_type": "multi_screenshot",
                "layout_index": 6,
            },
        }
        first_idx = min(funnel_idx, spend_idx)
        second_idx = max(funnel_idx, spend_idx)
        ordered[first_idx] = merged
        ordered.pop(second_idx)
        _report(ctx, f"   Merged {funnel_s['id']} + {spend_s['id']} at position {first_idx}")

    return ordered


def step_build_deck(ctx):
    """Build the final PowerPoint from all_slides."""
    if not ctx["_deck_available"]:
        _report(ctx, "⚠️ Skipping PowerPoint — deck_builder not available")
        return ctx

    SC = ctx["_SlideContent"]
    DB = ctx["_DeckBuilder"]
    template = ctx["_DECK_CONFIG"]["template_path"]

    # Reorder analysis slides to match reference section flow
    all_slides = ctx["all_slides"]
    selected = [s for s in all_slides if s.get("include", True)]
    ordered = _reorder_analysis_slides(ctx, selected)

    # Wire preamble slots: find slides to splice into preamble before
    # converting to SlideContent (so we can remove them from analysis list)
    wire_swipes = None
    wire_spend = None
    wire_trend = None

    # Find the most recent month for A12 wiring
    a12_months = set()
    for s in ordered:
        sid = s["id"]
        if sid.startswith("A12 - "):
            parts = sid.replace("A12 - ", "").rsplit(" ", 1)
            if len(parts) == 2:
                a12_months.add(parts[0])

    if a12_months:
        sorted_months = sorted(
            a12_months, key=lambda m: pd.to_datetime(m, format="%b%y", errors="coerce")
        )
        first_month = sorted_months[0]
        swipes_id = f"A12 - {first_month} Swipes"
        spend_id = f"A12 - {first_month} Spend"

        for s in ordered:
            if s["id"] == swipes_id:
                wire_swipes = s
            elif s["id"] == spend_id:
                wire_spend = s
            elif s["id"] == "A13.5 - Responder Count Trend":
                wire_trend = s
    else:
        for s in ordered:
            if s["id"] == "A13.5 - Responder Count Trend":
                wire_trend = s

    # Remove wired slides from analysis list (they go into preamble instead)
    wired_ids = set()
    if wire_swipes:
        wired_ids.add(wire_swipes["id"])
    if wire_spend:
        wired_ids.add(wire_spend["id"])
    if wire_trend:
        wired_ids.add(wire_trend["id"])

    # Also remove the first month's orphaned A13 summary — its A12 charts
    # are wired to preamble, so the summary alone in appendix is meaningless
    if a12_months:
        first_month_a13_id = f"A13 - {first_month} Mailer Summary"
        wired_ids.add(first_month_a13_id)

    ordered = [s for s in ordered if s["id"] not in wired_ids]

    # Convert analysis insight dicts → SlideContent objects
    # Section markers (dividers/placeholders) are converted directly
    analysis_slides = []
    for s in ordered:
        if s.get("__section__"):
            sc = SC(
                slide_type=s["slide_type"],
                title=s["title"],
                layout_index=s["layout_index"],
                kpis=s.get("kpis", {}),
            )
            analysis_slides.append(sc)
        else:
            sc = build_slide_content(ctx, s["data"])
            if sc:
                analysis_slides.append(sc)

    # Build preamble slides (always — even if no analysis yet)
    preamble = _build_preamble_slides(ctx)

    # Splice wired slides into preamble placeholder positions
    # P08 = index 7 (Swipes), P09 = index 8 (Spend), P13 = index 12 (Trend)
    if wire_swipes:
        sc = build_slide_content(ctx, wire_swipes["data"])
        if sc:
            sc.layout_index = 13
            preamble[7] = sc
            _report(ctx, f"   Wired P08 <- {wire_swipes['id']}")
    if wire_spend:
        sc = build_slide_content(ctx, wire_spend["data"])
        if sc:
            sc.layout_index = 13
            preamble[8] = sc
            _report(ctx, f"   Wired P09 <- {wire_spend['id']}")
    if wire_trend:
        trend_data = dict(wire_trend["data"])
        # Move original title to subtitle as insight text
        orig_title = trend_data.get("title", "")
        trend_data["title"] = f"Program Responses to Date\n{orig_title}"
        trend_data.pop("subtitle", None)
        sc = build_slide_content(ctx, trend_data)
        if sc:
            sc.layout_index = 13
            preamble[12] = sc
            _report(ctx, f"   Wired P13 <- {wire_trend['id']}")

    final_slides = preamble + analysis_slides

    if not final_slides:
        _report(ctx, "⚠️ No slides to build")
        return ctx

    # Build!
    pptx_path = ctx["pptx_output_path"]

    # Pre-flight: detect locked output file (open in PowerPoint)
    if os.path.exists(pptx_path):
        lock_file = Path(pptx_path).parent / f"~${Path(pptx_path).name}"
        if lock_file.exists():
            _report(
                ctx,
                f"❌ Cannot write — {os.path.basename(pptx_path)} is open in another application. Close it and retry.",
            )
            ctx["pptx_built"] = False
            return ctx

    _report(ctx, f"📽️ Building PowerPoint — {len(final_slides)} slides...")

    # Debug: check all layout indices before building
    from pptx import Presentation as _Prs

    _test_prs = _Prs(template)
    _max_layout = len(_test_prs.slide_layouts) - 1
    for i, s in enumerate(final_slides):
        if s.layout_index > _max_layout:
            print(
                f"🔴 BAD LAYOUT: slide {i} '{s.title}' type='{s.slide_type}' layout_index={s.layout_index} (max={_max_layout})"
            )
        else:
            print(f"   ✅ slide {i} '{s.title}' type='{s.slide_type}' layout={s.layout_index}")
    del _test_prs

    try:
        builder = DB(template)
        builder.build(final_slides, pptx_path)
        ctx["pptx_built"] = True
        _report(ctx, f"✅ PowerPoint saved: {os.path.basename(pptx_path)}")

        # Archive copy
        archive_path = ctx["pptx_archive_path"]
        Path(archive_path).parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(pptx_path, archive_path)
        _report(ctx, f"📁 Archived: {archive_path}")

    except Exception as e:
        _report(ctx, f"❌ PowerPoint build failed: {e}")
        ctx["pptx_built"] = False

    return ctx


# =============================================================================
# Archive Excel copy
# =============================================================================


def step_archive_excel(ctx):
    """Copy source Excel to archive folder."""
    src = ctx["source_excel_path"]
    dst = ctx["monthly_archive_file_path"]
    try:
        if os.path.exists(src):
            Path(dst).parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(src, dst)
            _report(ctx, f"📁 Excel archived: {dst}")
    except Exception as e:
        _report(ctx, f"⚠️ Archive copy failed: {e}")
    return ctx


# =============================================================================
# ANALYSES A1-A5 (same logic, updated to use source paths)
# =============================================================================


def run_a1(ctx):
    """A1: Stat Code Distribution."""
    _report(ctx, "\n🔍 A1 — Stat Code Distribution")
    data = ctx["data"]
    chart_dir = ctx["chart_dir"]

    data_clean = data.copy()
    data_clean["Stat Code"] = data_clean["Stat Code"].fillna("Unknown")
    data_clean["Business?"] = data_clean["Business?"].fillna("Unknown")

    grouped = data_clean.groupby(["Stat Code", "Business?"]).size().reset_index(name="Total Count")
    total = grouped["Total Count"].sum()

    output_data, summary_data = [], []
    label_map = {
        "Yes": "Business",
        "No": "Personal",
        "Y": "Business",
        "N": "Personal",
        "": "Unknown",
        "Unknown": "Unknown",
    }

    for sc in grouped["Stat Code"].unique():
        rows = grouped[grouped["Stat Code"] == sc]
        st = rows["Total Count"].sum()
        output_data.append(
            {
                "Stat Code": sc,
                "Account Type": "All",
                "Total Count": st,
                "Percent of Stat": st / total,
            }
        )
        biz, pers = 0, 0
        for _, r in rows.iterrows():
            lbl = label_map.get(str(r["Business?"]).strip(), str(r["Business?"]))
            cnt = r["Total Count"]
            if lbl == "Business":
                biz = cnt
            elif lbl == "Personal":
                pers = cnt
            output_data.append(
                {
                    "Stat Code": sc,
                    "Account Type": f"  → {lbl}",
                    "Total Count": cnt,
                    "Percent of Stat": cnt / st,
                }
            )
        summary_data.append(
            {
                "Stat Code": sc,
                "Total Count": st,
                "Percent of Total": st / total,
                "Business Count": biz,
                "Personal Count": pers,
            }
        )

    a1_dist = pd.DataFrame(output_data).sort_values(["Stat Code", "Account Type"])
    a1_summary = pd.DataFrame(summary_data).sort_values("Total Count", ascending=False)

    save_to_excel(
        ctx,
        df={"Distribution": a1_dist, "Summary": a1_summary},
        sheet_name="A1-StatCode",
        analysis_title="Stat Code Distribution Analysis",
        key_metrics={
            "Total Accounts": f"{len(data):,}",
            "Stat Codes": len(a1_summary),
            "Top Code": a1_summary.iloc[0]["Stat Code"] if len(a1_summary) > 0 else "N/A",
        },
    )

    # Chart
    chart_path = None
    try:
        fig, ax = _get_figure(ctx, "single")
        dp = a1_summary.head(10).sort_values("Total Count", ascending=True)
        ax.barh(dp["Stat Code"].astype(str), dp["Total Count"], color="#2E86AB")
        ax.set_xlabel("Count")
        ax.set_title("Stat Code Distribution - Top 10")
        ax.xaxis.set_major_formatter(FuncFormatter(lambda x, p: f"{x:,.0f}"))
        for i, (c, p) in enumerate(zip(dp["Total Count"], dp["Percent of Total"])):
            ax.text(
                c + dp["Total Count"].max() * 0.01,
                i,
                f"{c:,.0f} ({p:.1%})",
                va="center",
                fontsize=9,
            )
        plt.tight_layout()
        for ax in fig.get_axes():
            for spine in ax.spines.values():
                spine.set_visible(False)
        chart_path = str(chart_dir / "a1_stat_code.png")
        fig.savefig(chart_path, dpi=150, bbox_inches="tight", facecolor="white")
        plt.close(fig)
    except Exception as e:
        _report(ctx, f"   ⚠️ Chart failed: {e}")

    # Insight
    ta = a1_summary["Total Count"].sum()
    tc = a1_summary.iloc[0]["Stat Code"]
    tp = a1_summary.iloc[0]["Percent of Total"]
    t3 = a1_summary.head(3)["Percent of Total"].sum()
    tp_cnt = a1_summary["Personal Count"].sum()
    tb_cnt = a1_summary["Business Count"].sum()

    insight = {
        "title": "Stat Code Distribution",
        "subtitle": f"{tc} accounts represent {tp:.0%} of portfolio — {len(a1_summary)} stat codes total",
        "kpis": {"Total Accounts": f"{ta:,}", "Stat Codes": f"{len(a1_summary)}"},
        "layout_index": 4,
        "insights": [
            f"Top stat code '{tc}': {tp:.1%}",
            f"Top 3: {t3:.1%}",
            f"Personal: {tp_cnt:,}  |  Business: {tb_cnt:,}",
        ],
        "chart_path": chart_path,
        "category": "Overview",
    }
    if chart_path:
        add_to_library(ctx, "A1 - Stat Code Distribution", insight, "Overview")

    ctx["results"]["a1"] = {"distribution": a1_dist, "summary": a1_summary, "insight": insight}
    _report(ctx, f"✅ A1 complete — {len(a1_summary)} stat codes, top: {tc} ({tp:.1%})")
    return ctx


def run_a2(ctx):
    """A2: Debit Card Analysis."""
    _report(ctx, "\n💳 A2 — Debit Card Analysis")
    data = ctx["data"]
    chart_dir = ctx["chart_dir"]
    em = (
        ctx["eligible_mailable"]
        if isinstance(ctx["eligible_mailable"], list)
        else [ctx["eligible_mailable"]]
    )

    total_all = len(data)
    eligible = data[
        (data["Stat Code"].isin(ctx["eligible_stat_code"]))
        & (data["Prod Code"].isin(ctx["eligible_prod_code"]))
        & (data["Mailable?"].isin(em))
        & data["Date Closed"].isna()
        & (data["Avg Bal"] != 0)
    ]
    te = len(eligible)
    if te == 0:
        _report(ctx, "   ⚠️ No eligible accounts")
        return ctx

    ec = eligible.copy()
    ec["Debit?"] = ec["Debit?"].astype(str).str.strip()
    ec["Business?"] = ec["Business?"].astype(str).str.strip()

    wd = len(ec[ec["Debit?"] == "Yes"])
    wod = len(ec[ec["Debit?"] == "No"])
    dc = ec[ec["Debit?"] == "Yes"]
    pwd = len(dc[dc["Business?"] == "No"])
    bwd = len(dc[dc["Business?"] == "Yes"])

    rows = [
        {"Category": "Total Accounts (All)", "Count": total_all, "Percentage": 1.0},
        {"Category": "Total Eligible Accounts", "Count": te, "Percentage": te / total_all},
        {"Category": "→ Eligible with Debit Card", "Count": wd, "Percentage": wd / te},
        {"Category": "→ Eligible without Debit Card", "Count": wod, "Percentage": wod / te},
        {"Category": "  • Personal w/ Debit", "Count": pwd, "Percentage": pwd / wd if wd else 0},
        {"Category": "  • Business w/ Debit", "Count": bwd, "Percentage": bwd / wd if wd else 0},
    ]
    a2_summary = pd.DataFrame(rows)

    di = {
        "eligible_pct": te / total_all * 100,
        "penetration": wd / te * 100,
        "personal_pct": pwd / wd * 100 if wd else 0,
        "business_pct": bwd / wd * 100 if wd else 0,
    }

    save_to_excel(
        ctx,
        df=a2_summary,
        sheet_name="A2-DebitCard",
        analysis_title="Debit Card Summary Analysis",
        key_metrics={
            "Total": f"{total_all:,}",
            "Eligible": f"{di['eligible_pct']:.1f}%",
            "Penetration": f"{di['penetration']:.1f}%",
        },
    )

    chart_path = None
    try:
        fig, ax = _get_figure(ctx, "single")
        pd2 = a2_summary[~a2_summary["Category"].str.startswith("  ")].iloc[::-1]
        ax.barh(pd2["Category"], pd2["Count"], color="#2E86AB")
        ax.set_xlabel("Count")
        ax.xaxis.set_major_formatter(FuncFormatter(lambda x, p: f"{x:,.0f}"))
        for i, (c, p) in enumerate(zip(pd2["Count"], pd2["Percentage"])):
            ax.text(
                c + pd2["Count"].max() * 0.01, i, f"{c:,.0f} ({p:.1%})", va="center", fontsize=9
            )
        plt.tight_layout()
        for ax in fig.get_axes():
            for spine in ax.spines.values():
                spine.set_visible(False)
        chart_path = str(chart_dir / "a2_debit_card.png")
        fig.savefig(chart_path, dpi=150, bbox_inches="tight", facecolor="white")
        plt.close(fig)
    except Exception as e:
        _report(ctx, f"   ⚠️ Chart failed: {e}")

    insight = {
        "title": "Debit Card Analysis",
        "subtitle": f"{di['penetration']:.1f}% debit card penetration among eligible accounts",
        "kpis": {
            "Penetration": f"{di['penetration']:.1f}%",
            "Eligible": f"{di['eligible_pct']:.1f}%",
        },
        "layout_index": 5,
        "chart_path": chart_path,
        "category": "Overview",
    }
    if chart_path:
        add_to_library(ctx, "A2 - Debit Card Analysis", insight, "Overview")

    ctx["results"]["a2"] = {"summary": a2_summary, "insights": di, "insight": insight}
    _report(ctx, f"✅ A2 complete — {di['penetration']:.1f}% penetration")
    return ctx


def run_a3(ctx):
    """A3: Eligibility Funnel."""
    _report(ctx, "\n🎯 A3 — Eligibility Funnel")
    data = ctx["data"]
    chart_dir = ctx["chart_dir"]
    oa = ctx["open_accounts"]
    ed = ctx["eligible_data"]
    ep = ctx["eligible_personal"]
    eb = ctx["eligible_business"]
    esc = ctx["eligible_stat_code"]
    epc = ctx["eligible_prod_code"]
    em = (
        ctx["eligible_mailable"]
        if isinstance(ctx["eligible_mailable"], list)
        else [ctx["eligible_mailable"]]
    )

    ta = len(data)
    oc = len(oa)
    sc = len(oa[oa["Stat Code"].isin(esc)])
    se = oa[oa["Stat Code"].isin(esc)]
    pc = len(se[se["Prod Code"].isin(epc)])
    pe = se[se["Prod Code"].isin(epc)]
    mc = len(pe[pe["Mailable?"].isin(em)])
    ec = len(ed)

    funnel = pd.DataFrame(
        [
            {
                "Stage": "1. Total Accounts",
                "Count": ta,
                "Pct of Total": 1.0,
                "Drop-off": 0,
                "Drop-off %": 0.0,
            },
            {
                "Stage": "2. Open Accounts",
                "Count": oc,
                "Pct of Total": oc / ta,
                "Drop-off": ta - oc,
                "Drop-off %": (ta - oc) / ta,
            },
            {
                "Stage": "3. + Eligible Status Code",
                "Count": sc,
                "Pct of Total": sc / ta,
                "Drop-off": oc - sc,
                "Drop-off %": (oc - sc) / oc if oc else 0,
            },
            {
                "Stage": "4. + Eligible Product Code",
                "Count": pc,
                "Pct of Total": pc / ta,
                "Drop-off": sc - pc,
                "Drop-off %": (sc - pc) / sc if sc else 0,
            },
            {
                "Stage": "5. + Mailable",
                "Count": mc,
                "Pct of Total": mc / ta,
                "Drop-off": pc - mc,
                "Drop-off %": (pc - mc) / pc if pc else 0,
            },
            {
                "Stage": "6. + Has Balance (ELIGIBLE)",
                "Count": ec,
                "Pct of Total": ec / ta,
                "Drop-off": mc - ec,
                "Drop-off %": (mc - ec) / mc if mc else 0,
            },
        ]
    )
    if ec > 0:
        funnel = pd.concat(
            [
                funnel,
                pd.DataFrame(
                    [
                        {
                            "Stage": "   → Personal",
                            "Count": len(ep),
                            "Pct of Total": len(ep) / ec,
                            "Drop-off": None,
                            "Drop-off %": None,
                        },
                        {
                            "Stage": "   → Business",
                            "Count": len(eb),
                            "Pct of Total": len(eb) / ec,
                            "Drop-off": None,
                            "Drop-off %": None,
                        },
                    ]
                ),
            ],
            ignore_index=True,
        )

    drop_data = funnel[(funnel["Drop-off %"].notna()) & (funnel["Drop-off %"] > 0)]
    biggest = funnel.loc[drop_data["Drop-off %"].idxmax(), "Stage"] if len(drop_data) > 0 else "N/A"
    er = (ec / ta) * 100

    save_to_excel(
        ctx,
        df=funnel,
        sheet_name="A3-Eligibility",
        analysis_title="Eligibility Funnel",
        key_metrics={
            "Total": f"{ta:,}",
            "Eligible": f"{ec:,} ({er:.1f}%)",
            "Biggest Drop": biggest,
        },
    )

    # Table chart
    chart_path = None
    try:
        fig, ax = _get_figure(ctx, "single")
        ax.axis("off")
        tdf = funnel[~funnel["Stage"].str.startswith("   ")].copy()
        tdf["Count"] = tdf["Count"].apply(lambda x: f"{x:,}")
        tdf["Pct of Total"] = tdf["Pct of Total"].apply(lambda x: f"{x:.1%}")
        tdf["Drop-off"] = tdf["Drop-off"].apply(lambda x: f"{x:,}" if pd.notna(x) else "—")
        tdf["Drop-off %"] = tdf["Drop-off %"].apply(
            lambda x: f"{x:.1%}" if pd.notna(x) and x > 0 else "—"
        )
        t = ax.table(
            cellText=tdf.values,
            colLabels=tdf.columns,
            cellLoc="center",
            loc="center",
            colColours=["#2E86AB"] * len(tdf.columns),
        )
        t.auto_set_font_size(False)
        t.set_fontsize(9)
        t.scale(1.2, 1.8)
        for j in range(len(tdf.columns)):
            t[(0, j)].set_text_props(weight="bold", color="white")
        plt.tight_layout()
        for ax in fig.get_axes():
            for spine in ax.spines.values():
                spine.set_visible(False)
        chart_path = str(chart_dir / "a3_eligibility_funnel.png")
        fig.savefig(chart_path, dpi=150, bbox_inches="tight", facecolor="white")
        plt.close(fig)
    except Exception as e:
        _report(ctx, f"   ⚠️ Chart failed: {e}")

    bc = biggest.split(". ")[1] if ". " in biggest else biggest
    insight = {
        "title": "Eligibility Funnel Analysis",
        "subtitle": f"{er:.0f}% eligible — biggest drop-off at {bc}",
        "layout_index": 9,
        "kpis": {"Eligible": f"{er:.1f}%", "Accounts": f"{ec:,}"},
        "chart_path": chart_path,
        "category": "Overview",
    }
    if chart_path:
        add_to_library(ctx, "A3 - Eligibility Funnel", insight, "Overview")

    ctx["results"]["a3"] = {
        "funnel": funnel,
        "insight": insight,
        "insights": {
            "total_accounts": ta,
            "eligible_accounts": ec,
            "eligibility_rate": er,
            "biggest_dropoff": biggest,
            "personal_pct": len(ep) / ec * 100 if ec else 0,
            "business_pct": len(eb) / ec * 100 if ec else 0,
        },
    }
    _report(ctx, f"✅ A3 complete — {er:.1f}% eligible, biggest drop: {bc}")
    return ctx


def run_a4(ctx):
    """A4: Monthly Account Opening Trends."""
    _report(ctx, "\n📅 A4 — Monthly Trends")
    ed = ctx["eligible_data"]
    chart_dir = ctx["chart_dir"]

    if "Date Opened" not in ed.columns:
        _report(ctx, "   ⚠️ No Date Opened column")
        return ctx

    ed_copy = ed.copy()
    ed_copy["Date Opened"] = pd.to_datetime(ed_copy["Date Opened"], errors="coerce")
    valid = ed_copy[ed_copy["Date Opened"].notna()]
    if valid.empty:
        _report(ctx, "   ⚠️ No valid dates")
        return ctx

    valid["Year"] = valid["Date Opened"].dt.year
    monthly = valid.pivot_table(
        index="Year", columns=valid["Date Opened"].dt.month, aggfunc="size", fill_value=0
    )
    mn = {
        1: "Jan",
        2: "Feb",
        3: "Mar",
        4: "Apr",
        5: "May",
        6: "Jun",
        7: "Jul",
        8: "Aug",
        9: "Sep",
        10: "Oct",
        11: "Nov",
        12: "Dec",
    }
    monthly.columns = [mn.get(m, m) for m in monthly.columns]
    monthly["Year Total"] = monthly.sum(axis=1)
    monthly.index.name = "Year"
    trends = monthly.reset_index()
    trends["Year"] = trends["Year"].astype(int).astype(str)

    # Total row
    total_row = {"Year": "TOTAL"}
    total_row.update(trends.select_dtypes(include=[np.number]).sum().to_dict())
    trends = pd.concat([trends, pd.DataFrame([total_row])], ignore_index=True)

    gt = int(trends[trends["Year"] == "TOTAL"]["Year Total"].iloc[0])
    ny = len(trends) - 1

    save_to_excel(
        ctx,
        df=trends,
        sheet_name="A4-Trends",
        analysis_title="Monthly Account Opening Trends",
        key_metrics={"Total Openings": f"{gt:,}", "Years": f"{ny}"},
    )

    chart_path = None
    try:
        fig, ax = _get_figure(ctx, "single")
        dp = trends[trends["Year"] != "TOTAL"]
        ax.bar(dp["Year"], dp["Year Total"], color="#2E86AB")
        ax.set_ylabel("Openings")
        ax.set_title("Account Openings by Year")
        ax.yaxis.set_major_formatter(FuncFormatter(lambda x, p: f"{x:,.0f}"))
        for i, (y, c) in enumerate(zip(dp["Year"], dp["Year Total"])):
            ax.text(i, c + dp["Year Total"].max() * 0.02, f"{c:,.0f}", ha="center", fontsize=9)
        if ny > 10:
            plt.xticks(rotation=45)
        plt.tight_layout()
        for ax in fig.get_axes():
            for spine in ax.spines.values():
                spine.set_visible(False)
        chart_path = str(chart_dir / "a4_yearly_trends.png")
        fig.savefig(chart_path, dpi=150, bbox_inches="tight", facecolor="white")
        plt.close(fig)
    except Exception as e:
        _report(ctx, f"   ⚠️ Chart failed: {e}")

    insight = {
        "title": "Account Opening Trends",
        "subtitle": f"{gt:,} eligible accounts across {ny} years",
        "layout_index": 9,
        "kpis": {"Openings": f"{gt:,}", "Years": f"{ny}"},
        "chart_path": chart_path,
        "category": "Overview",
    }
    if chart_path:
        add_to_library(ctx, "A4 - Account Opening Trends", insight, "Overview")

    ctx["results"]["a4"] = {"trends": trends, "insight": insight}
    _report(ctx, f"✅ A4 complete — {gt:,} openings, {ny} years")
    return ctx


def run_a5(ctx):
    """A5: Branch Distribution."""
    _report(ctx, "\n🏢 A5 — Branch Distribution")
    ed = ctx["eligible_data"]
    chart_dir = ctx["chart_dir"]
    bm = ctx["config"].get("BranchMapping", {})

    ec = ed.copy()
    ec["Branch Name"] = ec["Branch"].map(bm).fillna(ec["Branch"]) if bm else ec["Branch"]

    rows = []
    for bn in sorted(ec["Branch Name"].unique()):
        ba = ec[ec["Branch Name"] == bn]
        rows.append(
            {
                "Branch": bn,
                "Total": len(ba),
                "Personal": len(ba[ba["Business?"] == "No"]),
                "Business": len(ba[ba["Business?"] == "Yes"]),
                "Pct of Total": len(ba) / len(ed) if len(ed) else 0,
            }
        )

    a5 = pd.DataFrame(rows).sort_values("Total", ascending=False)
    total_row = {
        "Branch": "TOTAL",
        "Total": a5["Total"].sum(),
        "Personal": a5["Personal"].sum(),
        "Business": a5["Business"].sum(),
        "Pct of Total": 1.0,
    }
    a5 = pd.concat([a5, pd.DataFrame([total_row])], ignore_index=True)

    dr = a5[a5["Branch"] != "TOTAL"]
    nb = len(dr)
    lb = dr.iloc[0]["Branch"]
    lp = dr.iloc[0]["Pct of Total"] * 100

    save_to_excel(
        ctx,
        df=a5,
        sheet_name="A5-Branch-Dist",
        analysis_title="Branch Distribution",
        key_metrics={"Branches": nb, "Largest": f"{lb} ({lp:.1f}%)"},
    )

    chart_path = None
    try:
        fig, ax = _get_figure(ctx, "single")
        dp = dr.head(10).iloc[::-1]
        ax.barh(dp["Branch"].astype(str), dp["Total"], color="#2E86AB")
        ax.set_xlabel("Count")
        ax.set_title("Branch Distribution - Top 10")
        ax.xaxis.set_major_formatter(FuncFormatter(lambda x, p: f"{x:,.0f}"))
        for i, (c, p) in enumerate(zip(dp["Total"], dp["Pct of Total"])):
            ax.text(c + dp["Total"].max() * 0.01, i, f"{c:,.0f} ({p:.1%})", va="center", fontsize=9)
        plt.tight_layout()
        for ax in fig.get_axes():
            for spine in ax.spines.values():
                spine.set_visible(False)
        chart_path = str(chart_dir / "a5_branch_distribution.png")
        fig.savefig(chart_path, dpi=150, bbox_inches="tight", facecolor="white")
        plt.close(fig)
    except Exception as e:
        _report(ctx, f"   ⚠️ Chart failed: {e}")

    insight = {
        "title": "Branch Distribution",
        "subtitle": f"{lb} leads with {lp:.0f}% — {nb} branches total",
        "layout_index": 4,
        "kpis": {"Branches": f"{nb}", "Top": lb},
        "chart_path": chart_path,
        "category": "Overview",
    }
    if chart_path:
        add_to_library(ctx, "A5 - Branch Distribution", insight, "Overview")

    ctx["results"]["a5"] = {"distribution": a5, "insight": insight}
    _report(ctx, f"✅ A5 complete — {nb} branches, largest: {lb} ({lp:.1f}%)")
    return ctx


# =============================================================================
# MAIN PIPELINE RUNNER
# =============================================================================

# =============================================================================
# A9 — ATTRITION ANALYSIS
# =============================================================================


def run_a9(ctx):
    """A9 — Account Attrition Analysis.

    Delegates to attrition.run_attrition_suite which runs 13 sub-analyses,
    each producing a table + chart + slide in one pass.

    Outputs stored in ctx['results']['attrition_1'] .. ctx['results']['attrition_13'].
    """
    from ars_analysis.attrition import run_attrition_suite

    return run_attrition_suite(ctx)


# =============================================================================
# A6 — DCTR SUITE  (consolidated A6 data + A7 charts)
# =============================================================================


def run_a6(ctx):
    """A6 — Comprehensive Debit Card Take Rate analysis.

    Delegates to dctr.run_dctr_suite which runs 16 sub-analyses,
    each producing a table + chart + slide in one pass.

    Outputs stored in ctx['results']['dctr_1'] .. ctx['results']['dctr_16'].
    See DCTR_DEPENDENCY_MAP.md for full dependency documentation.
    """
    from ars_analysis.dctr import run_dctr_suite

    return run_dctr_suite(ctx)


def run_a8(ctx):
    """A8 — Comprehensive Reg E (Regulation E) opt-in analysis.

    Delegates to reg_e.run_reg_e_suite which runs 13 sub-analyses,
    each producing a table + chart + slide in one pass.

    Outputs stored in ctx['results']['reg_e_1'] .. ctx['results']['reg_e_13'].
    """
    from ars_analysis.reg_e import run_reg_e_suite

    return run_reg_e_suite(ctx)


def run_a11(ctx):
    """A11 — Value Analysis (debit card + Reg E).

    Delegates to value.run_value_suite which runs A11.1 and A11.2.
    Outputs stored in ctx['results']['value_1'] and ctx['results']['value_2'].
    """
    from ars_analysis.value import run_value_suite

    return run_value_suite(ctx)


def run_a12(ctx):
    """A12 — Mail Campaign Insights (Spend & Swipes per month).

    Delegates to mailer_insights.run_mailer_insights_suite.
    Produces 2 slides per mail month (Spend + Swipes).
    """
    from ars_analysis.mailer_insights import run_mailer_insights_suite

    return run_mailer_insights_suite(ctx)


def run_a13_14(ctx):
    """A13+A14 — Mailer Response & Demographics.

    Delegates to mailer_response.run_mailer_response_suite.
    A13: Response summary, segment perf, counts, share, trend.
    A14: Response by segment, responder account age.
    """
    from ars_analysis.mailer_response import run_mailer_response_suite

    return run_mailer_response_suite(ctx)


def run_a15(ctx):
    """A15 — Market Impact (reach bubble + spend share + revenue + delta).

    Delegates to mailer_impact.run_mailer_impact_suite.
    A15.1: Market reach — eligible w/ card vs unique responders.
    A15.2: Spend share — total spend across open, eligible, responders.
    A15.3: Revenue attribution — IC revenue from responders vs non-responders.
    A15.4: Pre/post spend delta — before vs after mailer comparison.
    """
    from ars_analysis.mailer_impact import run_mailer_impact_suite

    return run_mailer_impact_suite(ctx)


def run_pipeline(
    file_path,
    config_path=None,
    base_paths=None,
    deck_builder_path=None,
    template_path=None,
    progress_callback=None,
    exec_report=None,
    metadata=None,
    deck_only=False,
):
    """
    Run the full ARS pipeline: Cells 1-16 + A1-A5 + Build Deck.

    Set deck_only=True to skip Excel writes and archive -- just regenerate
    charts and rebuild the PowerPoint.  Much faster for iterating on
    slide layout and chart design.

    Returns the pipeline context with all results.
    """
    ctx = create_context()
    if deck_only:
        ctx["_skip_excel"] = True
    if progress_callback:
        ctx["_progress_callback"] = progress_callback
    if exec_report:
        ctx["_exec_report"] = exec_report

    # Verify we're running from the project venv
    import sys

    _report(ctx, "=" * 60)
    _report(ctx, "🚀 ARS PIPELINE — STARTING")
    _report(ctx, f"   Python: {sys.executable}")
    if ".venv" not in sys.executable:
        _report(ctx, "")
        _report(ctx, "⚠️  WARNING: You are NOT running from the project .venv!")
        _report(ctx, "   Your code changes will NOT take effect.")
        _report(ctx, "   Fix: use  uv run python -m ars_analysis.pipeline ...")
        _report(ctx, "   Or:  .venv\\Scripts\\activate  then  python -m ...")
        _report(ctx, "")
    _report(ctx, "=" * 60)
    start_time = time.time()
    ctx["_phase_start"] = start_time

    # Pre-analysis (Cells 1-16)
    _report(ctx, "\n📦 Phase 1: Setup")
    _report(ctx, "   Loading data file...")
    ctx = step_load_data(ctx, file_path)
    _report(ctx, "   Parsing filename & paths...")
    ctx = step_parse_and_paths(ctx, base_paths, metadata=metadata)
    _report(ctx, "   Loading client config...")
    ctx = step_load_config(ctx, config_path)
    _report(ctx, "   Cleaning data...")
    ctx = step_clean_data(ctx)
    _report(ctx, "   Calculating date range...")
    ctx = step_date_range(ctx)
    _report(ctx, "   Creating data subsets...")
    ctx = step_create_subsets(ctx)
    _report(ctx, "   Setting up deck builder...")
    ctx = step_setup_deck(ctx, deck_builder_path, template_path)
    _report(ctx, "   ✅ Setup complete")
    _exec_report(ctx, "set_portfolio")
    _phase_time(ctx, "Setup")

    # Analyses A1-A6
    _report(ctx, "\n📊 Phase 2: Analyses (A1-A5)")
    for run_fn, label in [
        (run_a1, "A1"),
        (run_a2, "A2"),
        (run_a3, "A3"),
        (run_a4, "A4"),
        (run_a5, "A5"),
    ]:
        try:
            ctx = run_fn(ctx)
        except Exception as e:
            _report(ctx, f"   ⚠️ {label} failed: {e}")
            import traceback

            traceback.print_exc()
    _phase_time(ctx, "A1-A5")
    _exec_report(ctx, "set_overview")

    _report(ctx, "\n📉 Phase 2b: Attrition Analysis (A9)")
    try:
        ctx = run_a9(ctx)
        _report(ctx, "   ✅ Attrition suite complete")
        _exec_report(ctx, "set_attrition")
    except Exception as e:
        _report(ctx, f"   ⚠️ Attrition suite failed: {e}")
        _exec_report_fail(ctx, "attrition")
        import traceback

        traceback.print_exc()
    _phase_time(ctx, "Attrition")

    _report(ctx, "\n💳 Phase 3: DCTR Analysis (A6+A7)")
    try:
        ctx = run_a6(ctx)
        _report(ctx, "   ✅ DCTR suite complete")
        _exec_report(ctx, "set_dctr")
    except Exception as e:
        _report(ctx, f"   ⚠️ DCTR suite failed: {e}")
        _exec_report_fail(ctx, "dctr")
        import traceback

        traceback.print_exc()
    _phase_time(ctx, "DCTR")

    _report(ctx, "\n📋 Phase 3b: Reg E Analysis (A8)")
    try:
        ctx = run_a8(ctx)
        _report(ctx, "   ✅ Reg E suite complete")
        _exec_report(ctx, "set_reg_e")
    except Exception as e:
        _report(ctx, f"   ⚠️ Reg E suite failed: {e}")
        _exec_report_fail(ctx, "reg_e")
        import traceback

        traceback.print_exc()
    _phase_time(ctx, "Reg E")

    _report(ctx, "\n💰 Phase 3c: Value Analysis (A11)")
    try:
        ctx = run_a11(ctx)
        _report(ctx, "   ✅ Value suite complete")
        _exec_report(ctx, "set_value")
    except Exception as e:
        _report(ctx, f"   ⚠️ Value suite failed: {e}")
        _exec_report_fail(ctx, "value")
        import traceback

        traceback.print_exc()
    _phase_time(ctx, "Value")

    _report(ctx, "\n📧 Phase 3d: Mailer Insights (A12)")
    try:
        ctx = run_a12(ctx)
        _report(ctx, "   ✅ Mailer Insights suite complete")
        _exec_report(ctx, "set_mailer_insights")
    except Exception as e:
        _report(ctx, f"   ⚠️ Mailer Insights suite failed: {e}")
        _exec_report_fail(ctx, "mailer_insights")
        import traceback

        traceback.print_exc()
    _phase_time(ctx, "Mailer")

    _report(ctx, "\n📧 Phase 3e: Mailer Response & Demographics (A13+A14)")
    try:
        ctx = run_a13_14(ctx)
        _report(ctx, "   ✅ Mailer Response suite complete")
        _exec_report(ctx, "set_mailer_response")
    except Exception as e:
        _report(ctx, f"   ⚠️ Mailer Response suite failed: {e}")
        _exec_report_fail(ctx, "mailer_response")
        import traceback

        traceback.print_exc()
    _phase_time(ctx, "Response")

    _report(ctx, "\n📊 Phase 3f: Market Impact (A15)")
    try:
        ctx = run_a15(ctx)
        _report(ctx, "   ✅ Market Impact suite complete")
        _exec_report(ctx, "set_market_impact")
    except Exception as e:
        _report(ctx, f"   ⚠️ Market Impact suite failed: {e}")
        _exec_report_fail(ctx, "market_impact")
        import traceback

        traceback.print_exc()
    _phase_time(ctx, "Impact")

    if not ctx.get("_skip_excel"):
        # Flush cached Excel workbooks to disk
        _report(ctx, "\n💾 Saving Excel workbooks...")
        flush_workbooks(ctx)
        _phase_time(ctx, "Excel Save")
    else:
        _report(ctx, "\n⏩ Skipping Excel (deck_only mode)")

    # Build deck
    _report(ctx, "\n📽️ Phase 4: PowerPoint")
    ctx = step_build_deck(ctx)
    _phase_time(ctx, "Deck Build")

    if not ctx.get("_skip_excel"):
        # Archive
        ctx = step_archive_excel(ctx)

    elapsed = time.time() - start_time
    _report(ctx, f"\n{'=' * 60}")
    _report(ctx, f"✅ PIPELINE COMPLETE — {elapsed:.1f}s")
    _report(ctx, f"   Client: {ctx['client_id']} — {ctx['client_name']}")
    _report(ctx, f"   Records: {len(ctx['data']):,}")
    _report(ctx, f"   Exports: {len(ctx['export_log'])}")
    _report(ctx, f"   Slides:  {len(ctx['all_slides'])}")
    _report(ctx, f"   PPTX:    {'✅ Built' if ctx['pptx_built'] else '⬜ Skipped'}")
    _report(ctx, f"{'=' * 60}")
    _exec_report(ctx, "set_build_summary")

    return ctx


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python pipeline.py <path_to_odd_file.xlsx> [config_path]")
        sys.exit(1)
    run_pipeline(sys.argv[1], config_path=sys.argv[2] if len(sys.argv) > 2 else None)
