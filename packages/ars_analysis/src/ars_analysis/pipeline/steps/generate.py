"""Step: Generate deliverables (Excel, PowerPoint, archive)."""

from __future__ import annotations

import shutil

import openpyxl
from loguru import logger

from ars_analysis.output.excel_formatter import (
    create_summary_sheet,
    format_worksheet,
)
from ars_analysis.pipeline.context import PipelineContext


def step_generate(ctx: PipelineContext) -> None:
    """Generate all output deliverables from analysis results.

    Order: Excel workbook -> PowerPoint deck -> archive copy.
    Uses single-write pattern: build Excel once, then shutil.copy2 for master.
    """
    if not ctx.all_slides:
        logger.warning("No analysis results to generate deliverables from")
        return

    _write_excel(ctx)
    _build_deck(ctx)
    logger.info("Deliverables generated for {client}", client=ctx.client.client_id)


def step_archive(ctx: PipelineContext) -> None:
    """Copy deliverables to archive location. Non-critical step."""
    logger.info("Archive step for {client} (not yet implemented)", client=ctx.client.client_id)


def _write_excel(ctx: PipelineContext) -> None:
    """Write all analysis results to a formatted Excel workbook.

    Single-write pattern: one workbook with a tab per analysis.
    Then shutil.copy2 for the master/archive copy.
    """
    excel_path = ctx.paths.excel_dir / f"{ctx.client.client_id}_{ctx.client.month}_analysis.xlsx"
    ctx.paths.excel_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    sheets_written = 0
    for result in ctx.all_slides:
        if result.excel_data is None:
            continue
        for sheet_name, df in result.excel_data.items():
            # Truncate sheet name to Excel 31-char limit
            safe_name = f"{result.slide_id}_{sheet_name}"[:31]
            ws = wb.create_sheet(title=safe_name)
            # Write headers
            for col_idx, col_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
            # Write data rows
            for row_idx, row in enumerate(df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            # Format the worksheet
            format_worksheet(ws)
            sheets_written += 1

    if sheets_written == 0:
        logger.warning("No Excel data to write")
        return

    # Add summary sheet at the front
    create_summary_sheet(wb, ctx)

    wb.save(excel_path)
    ctx.export_log.append(str(excel_path))
    logger.info("Excel written: {path} ({n} sheets)", path=excel_path.name, n=sheets_written)

    # Single-write pattern: copy to master location if configured
    if ctx.settings and hasattr(ctx.settings, "paths"):
        master_dir = getattr(ctx.settings.paths, "ars_base", None)
        if master_dir and master_dir != ctx.paths.excel_dir:
            master_path = master_dir / excel_path.name
            try:
                shutil.copy2(excel_path, master_path)
                logger.info("Master copy: {name}", name=master_path.name)
            except OSError as exc:
                logger.warning("Master copy failed: {err}", err=exc)


def _build_deck(ctx: PipelineContext) -> None:
    """Build PowerPoint deck from analysis results."""
    skip_pptx = False
    if ctx.settings and hasattr(ctx.settings, "pipeline"):
        skip_pptx = getattr(ctx.settings.pipeline, "skip_pptx", False)

    if skip_pptx:
        logger.info("PowerPoint generation skipped (skip_pptx=True)")
        return

    try:
        from ars_analysis.output.deck_builder import build_deck

        build_deck(ctx)
    except ImportError:
        logger.info(
            "Deck build: {n} slides ready (deck_builder not available)",
            n=len(ctx.all_slides),
        )
    except Exception as exc:
        logger.warning("Deck build failed: {err}", err=exc)
