"""Results Viewer page -- browse outputs from any pipeline run."""

from __future__ import annotations

import io
import logging
import zipfile
from pathlib import Path

import streamlit as st

from platform_app.components import kpi_row

logger = logging.getLogger(__name__)


def _get_page(name: str):
    return st.session_state.get("_pages", {}).get(name)


def render() -> None:
    try:
        _render_inner()
    except Exception:
        st.error("Something went wrong loading results. Please try refreshing.")
        logger.exception("Unhandled error in results_viewer")


def _render_inner() -> None:
    st.title("Results Viewer")
    st.caption("View and download analysis results from any pipeline")

    # Show most recent run results
    history = st.session_state.get("run_history", [])
    successful_runs = [r for r in history if r.get("success") and r.get("output_dir")]

    if not successful_runs:
        st.info("No results to display yet. Run an analysis to see results here.")
        run_pg = _get_page("run")
        if run_pg and st.button("Go to Run Analysis", type="primary"):
            st.switch_page(run_pg)
        return

    # Run selector
    run_options = []
    for r in reversed(successful_runs[-15:]):
        label = (
            f"{r.get('pipeline', '').upper()} | "
            f"{r.get('client_id', '')} -- {r.get('client_name', '')} | "
            f"{r.get('timestamp', '')[:16]}"
        )
        run_options.append((label, r))

    selected_label = st.selectbox(
        "Select a run to view",
        [opt[0] for opt in run_options],
        index=0,
    )

    selected_run = next(r for label, r in run_options if label == selected_label)
    output_dir = Path(selected_run.get("output_dir", ""))

    if not output_dir.exists():
        st.warning("Output directory no longer available for this run.")
        return

    # KPI summary
    result_count = selected_run.get("result_count", 0)
    elapsed = selected_run.get("elapsed", 0)
    pipeline = selected_run.get("pipeline", "")

    kpi_row(
        [
            {"label": "Pipeline", "value": pipeline.upper()},
            {"label": "Analyses", "value": str(result_count)},
            {"label": "Time", "value": f"{elapsed:.1f}s"},
        ]
    )

    # File downloads
    st.markdown("### Downloads")
    downloadable = sorted(
        f
        for f in output_dir.rglob("*")
        if f.is_file() and f.suffix in (".xlsx", ".pptx", ".png", ".csv")
    )

    if not downloadable:
        st.caption("No output files found.")
        return

    # Group by type
    excel_files = [f for f in downloadable if f.suffix in (".xlsx",)]
    pptx_files = [f for f in downloadable if f.suffix == ".pptx"]
    chart_files = [f for f in downloadable if f.suffix == ".png"]
    csv_files = [f for f in downloadable if f.suffix == ".csv"]

    # Download buttons row
    col1, col2 = st.columns(2)
    with col1:
        for f in excel_files:
            st.download_button(
                f"Download {f.name}",
                f.read_bytes(),
                file_name=f.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
    with col2:
        for f in pptx_files:
            st.download_button(
                f"Download {f.name}",
                f.read_bytes(),
                file_name=f.name,
                mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                use_container_width=True,
            )

    # Charts as ZIP
    if chart_files:
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for chart in chart_files:
                zf.write(chart, chart.name)
        st.download_button(
            f"Export All Charts as ZIP ({len(chart_files)} charts)",
            data=buf.getvalue(),
            file_name=f"{selected_run.get('client_id', 'charts')}_charts.zip",
            mime="application/zip",
            use_container_width=True,
        )

    # Chart gallery
    if chart_files:
        st.markdown("### Charts")
        for i in range(0, len(chart_files), 2):
            cols = st.columns(2)
            for j, col in enumerate(cols):
                idx = i + j
                if idx < len(chart_files):
                    with col:
                        st.image(str(chart_files[idx]), use_container_width=True)
                        st.caption(chart_files[idx].stem)

    # Excel data preview
    if excel_files:
        st.markdown("### Data Preview")
        for excel_path in excel_files:
            with st.expander(excel_path.name, expanded=False):
                try:
                    import openpyxl
                    import pandas as pd

                    wb = openpyxl.load_workbook(excel_path, read_only=True)
                    tabs = st.tabs(wb.sheetnames)
                    for tab, sheet_name in zip(tabs, wb.sheetnames):
                        with tab:
                            df = pd.read_excel(excel_path, sheet_name=sheet_name)
                            st.dataframe(df, use_container_width=True, hide_index=True)
                except ImportError:
                    st.warning("Install openpyxl to preview Excel data.")
