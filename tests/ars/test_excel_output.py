"""Tests for Excel formatting and generation."""

import openpyxl
import pandas as pd

from ars_analysis.analytics.base import AnalysisResult
from ars_analysis.output.excel_formatter import (
    auto_column_width,
    create_summary_sheet,
    format_headers,
)
from ars_analysis.pipeline.context import ClientInfo, OutputPaths, PipelineContext
from ars_analysis.pipeline.steps.generate import _write_excel


def _make_ctx(tmp_path, with_slides=True):
    ctx = PipelineContext(
        client=ClientInfo(client_id="1234", client_name="Test CU", month="2026.01"),
        paths=OutputPaths.from_base(tmp_path, "1234", "2026.01"),
    )
    if with_slides:
        df = pd.DataFrame({"Account": ["A1", "A2"], "Balance": [100.0, 200.0]})
        ctx.all_slides = [
            AnalysisResult(
                slide_id="TEST-1",
                title="Test Analysis",
                excel_data={"test_data": df},
            )
        ]
    return ctx


class TestFormatHeaders:
    """format_headers applies styling to first row."""

    def test_bold_headers(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Header1")
        ws.cell(row=1, column=2, value="Header2")

        format_headers(ws)
        assert ws.cell(row=1, column=1).font.bold
        assert ws.freeze_panes == "A2"


class TestAutoColumnWidth:
    """auto_column_width sets reasonable widths."""

    def test_adjusts_width(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Short")
        ws.cell(row=2, column=1, value="A much longer cell value here")

        auto_column_width(ws)
        assert ws.column_dimensions["A"].width > 8


class TestCreateSummarySheet:
    """create_summary_sheet adds a Summary tab."""

    def test_creates_sheet(self, tmp_path):
        ctx = _make_ctx(tmp_path)
        wb = openpyxl.Workbook()
        create_summary_sheet(wb, ctx)
        assert "Summary" in wb.sheetnames

    def test_summary_content(self, tmp_path):
        ctx = _make_ctx(tmp_path)
        wb = openpyxl.Workbook()
        create_summary_sheet(wb, ctx)
        ws = wb["Summary"]
        assert ws.cell(row=1, column=1).value == "ARS Analysis Summary"
        assert ws.cell(row=4, column=2).value == "1234"  # Client ID


class TestWriteExcel:
    """_write_excel produces a formatted workbook."""

    def test_creates_file(self, tmp_path):
        ctx = _make_ctx(tmp_path)
        _write_excel(ctx)
        excel_path = ctx.paths.excel_dir / "1234_2026.01_analysis.xlsx"
        assert excel_path.exists()

    def test_has_summary_sheet(self, tmp_path):
        ctx = _make_ctx(tmp_path)
        _write_excel(ctx)
        excel_path = ctx.paths.excel_dir / "1234_2026.01_analysis.xlsx"
        wb = openpyxl.load_workbook(excel_path)
        assert "Summary" in wb.sheetnames

    def test_has_data_sheets(self, tmp_path):
        ctx = _make_ctx(tmp_path)
        _write_excel(ctx)
        excel_path = ctx.paths.excel_dir / "1234_2026.01_analysis.xlsx"
        wb = openpyxl.load_workbook(excel_path)
        assert len(wb.sheetnames) >= 2  # Summary + at least 1 data sheet

    def test_logged_to_export(self, tmp_path):
        ctx = _make_ctx(tmp_path)
        _write_excel(ctx)
        assert len(ctx.export_log) == 1

    def test_no_slides_no_file(self, tmp_path):
        ctx = _make_ctx(tmp_path, with_slides=False)
        _write_excel(ctx)
        assert len(ctx.export_log) == 0
