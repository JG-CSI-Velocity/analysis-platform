"""Tests for the generate step (Excel output + run report)."""

import json

import pandas as pd
import pytest

from ars_analysis.analytics.base import AnalysisResult
from ars_analysis.pipeline.context import ClientInfo, OutputPaths, PipelineContext
from ars_analysis.pipeline.steps.generate import (
    SlideStatus,
    _build_run_report,
    _save_run_report,
    step_generate,
)


@pytest.fixture
def ctx_with_results(tmp_path):
    """PipelineContext with analysis results ready for output."""
    paths = OutputPaths(
        base_dir=tmp_path,
        charts_dir=tmp_path / "charts",
        excel_dir=tmp_path,
        pptx_dir=tmp_path,
    )
    ctx = PipelineContext(
        client=ClientInfo(client_id="1200", client_name="Test CU", month="2026.02"),
        paths=paths,
    )
    ctx.all_slides = [
        AnalysisResult(
            slide_id="A1",
            title="Stat Codes",
            excel_data={
                "Summary": pd.DataFrame({"Code": ["O", "C"], "Count": [6, 3]}),
            },
        ),
        AnalysisResult(
            slide_id="A2",
            title="Products",
            excel_data={
                "Products": pd.DataFrame({"Product": ["DDA"], "Count": [5]}),
            },
        ),
    ]
    return ctx


def test_generate_writes_excel(ctx_with_results):
    step_generate(ctx_with_results)
    excel_files = list(ctx_with_results.paths.excel_dir.glob("*.xlsx"))
    assert len(excel_files) == 1
    assert "1200" in excel_files[0].name


def test_generate_excel_has_correct_sheets(ctx_with_results):
    step_generate(ctx_with_results)
    import openpyxl

    excel_path = list(ctx_with_results.paths.excel_dir.glob("*.xlsx"))[0]
    wb = openpyxl.load_workbook(excel_path)
    assert len(wb.sheetnames) == 3  # Summary, A1_Summary, A2_Products
    assert wb.sheetnames[0] == "Summary"


def test_generate_no_results_is_noop(tmp_path):
    ctx = PipelineContext(
        client=ClientInfo(client_id="1200", client_name="Test CU", month="2026.02"),
        paths=OutputPaths(base_dir=tmp_path, excel_dir=tmp_path),
    )
    step_generate(ctx)
    excel_files = list(tmp_path.glob("*.xlsx"))
    assert len(excel_files) == 0


def test_generate_logs_to_export_log(ctx_with_results):
    step_generate(ctx_with_results)
    assert len(ctx_with_results.export_log) >= 1  # Excel + possibly PPTX
    assert any("1200" in entry for entry in ctx_with_results.export_log)


def test_generate_result_without_excel_data(tmp_path):
    """Results with no excel_data should be skipped without error."""
    ctx = PipelineContext(
        client=ClientInfo(client_id="1200", client_name="Test CU", month="2026.02"),
        paths=OutputPaths(base_dir=tmp_path, excel_dir=tmp_path),
    )
    ctx.all_slides = [
        AnalysisResult(slide_id="A1", title="Chart Only", chart_path=None),
    ]
    step_generate(ctx)
    # No excel data means no file written
    excel_files = list(tmp_path.glob("*.xlsx"))
    assert len(excel_files) == 0


# -- Run Report tests --


class TestBuildRunReport:
    def test_basic_report_from_results(self, tmp_path):
        chart = tmp_path / "chart.png"
        chart.write_bytes(b"PNG")

        ctx = PipelineContext(
            client=ClientInfo(client_id="1200", client_name="Test CU", month="2026.02"),
            paths=OutputPaths(base_dir=tmp_path, excel_dir=tmp_path),
        )
        ctx.all_slides = [
            AnalysisResult(slide_id="A1", title="OK Slide", chart_path=chart, success=True),
            AnalysisResult(slide_id="A2", title="Failed", success=False, error="No data"),
        ]
        ctx.results["overview"] = [ctx.all_slides[0]]
        ctx.results["mailer"] = [ctx.all_slides[1]]

        report = _build_run_report(ctx)
        assert len(report) == 2
        assert report[0].slide_id == "A1"
        assert report[0].success is True
        assert report[0].has_chart is True
        assert report[0].module_id == "overview"
        assert report[1].slide_id == "A2"
        assert report[1].success is False
        assert report[1].error == "No data"
        assert report[1].module_id == "mailer"

    def test_report_missing_chart_file(self, tmp_path):
        ctx = PipelineContext(
            client=ClientInfo(client_id="1200", client_name="Test CU", month="2026.02"),
            paths=OutputPaths(base_dir=tmp_path, excel_dir=tmp_path),
        )
        ctx.all_slides = [
            AnalysisResult(
                slide_id="A1",
                title="Missing Chart",
                chart_path=tmp_path / "nonexistent.png",
                success=True,
            ),
        ]
        report = _build_run_report(ctx)
        assert report[0].has_chart is False

    def test_report_with_excel_data(self, tmp_path):
        ctx = PipelineContext(
            client=ClientInfo(client_id="1200", client_name="Test CU", month="2026.02"),
            paths=OutputPaths(base_dir=tmp_path, excel_dir=tmp_path),
        )
        ctx.all_slides = [
            AnalysisResult(
                slide_id="A1",
                title="With Excel",
                excel_data={"Sheet1": pd.DataFrame({"A": [1]})},
                success=True,
            ),
            AnalysisResult(slide_id="A2", title="No Excel", success=True),
        ]
        report = _build_run_report(ctx)
        assert report[0].has_excel is True
        assert report[1].has_excel is False

    def test_empty_slides(self, tmp_path):
        ctx = PipelineContext(
            client=ClientInfo(client_id="1200", client_name="Test CU", month="2026.02"),
            paths=OutputPaths(base_dir=tmp_path, excel_dir=tmp_path),
        )
        report = _build_run_report(ctx)
        assert report == []


class TestSaveRunReport:
    def test_saves_json(self, tmp_path):
        ctx = PipelineContext(
            client=ClientInfo(client_id="1200", client_name="Test CU", month="2026.02"),
            paths=OutputPaths(base_dir=tmp_path, excel_dir=tmp_path),
        )
        report = [
            SlideStatus(slide_id="A1", module_id="overview", success=True, has_chart=True, has_excel=True, title="OK"),
            SlideStatus(slide_id="A2", module_id="mailer", success=False, has_chart=False, has_excel=False, error="No data", title="Fail"),
        ]
        _save_run_report(ctx, report)

        report_path = tmp_path / "1200_2026.02_run_report.json"
        assert report_path.exists()

        data = json.loads(report_path.read_text())
        assert data["client_id"] == "1200"
        assert data["summary"]["total"] == 2
        assert data["summary"]["ok"] == 1
        assert data["summary"]["failed"] == 1
        assert len(data["slides"]) == 2

    def test_report_in_export_log(self, tmp_path):
        ctx = PipelineContext(
            client=ClientInfo(client_id="1200", client_name="Test CU", month="2026.02"),
            paths=OutputPaths(base_dir=tmp_path, excel_dir=tmp_path),
        )
        _save_run_report(ctx, [])
        assert any("run_report" in e for e in ctx.export_log)


class TestStepGenerateRunReport:
    def test_generate_creates_run_report(self, ctx_with_results):
        step_generate(ctx_with_results)
        report_files = list(ctx_with_results.paths.base_dir.glob("*_run_report.json"))
        assert len(report_files) == 1

    def test_run_report_stored_in_ctx_results(self, ctx_with_results):
        step_generate(ctx_with_results)
        assert "_run_report" in ctx_with_results.results
        assert len(ctx_with_results.results["_run_report"]) == 2

    def test_run_report_summary_counts(self, ctx_with_results):
        step_generate(ctx_with_results)
        report_path = list(ctx_with_results.paths.base_dir.glob("*_run_report.json"))[0]
        data = json.loads(report_path.read_text())
        assert data["summary"]["total"] == 2
