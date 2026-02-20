"""Tests for Excel report generation."""

from __future__ import annotations

import pytest
from openpyxl import load_workbook

from txn_analysis.exports.excel_report import write_excel_report
from txn_analysis.pipeline import PipelineResult, run_pipeline
from txn_analysis.settings import Settings


@pytest.fixture()
def pipeline_result(sample_csv_path, tmp_path):
    settings = Settings.from_args(data_file=sample_csv_path, output_dir=tmp_path)
    return run_pipeline(settings)


class TestWriteExcelReport:
    def test_creates_file(self, pipeline_result, tmp_path):
        path = tmp_path / "test_report.xlsx"
        write_excel_report(pipeline_result, path)
        assert path.exists()

    def test_has_cover_sheet(self, pipeline_result, tmp_path):
        path = tmp_path / "test_report.xlsx"
        write_excel_report(pipeline_result, path)
        wb = load_workbook(path)
        assert "Report Info" in wb.sheetnames

    def test_has_contents_sheet(self, pipeline_result, tmp_path):
        path = tmp_path / "test_report.xlsx"
        write_excel_report(pipeline_result, path)
        wb = load_workbook(path)
        assert "Contents" in wb.sheetnames

    def test_has_analysis_sheets(self, pipeline_result, tmp_path):
        path = tmp_path / "test_report.xlsx"
        write_excel_report(pipeline_result, path)
        wb = load_workbook(path)
        # Report Info + Contents + analysis sheets
        assert len(wb.sheetnames) > 2

    def test_cover_sheet_has_title(self, pipeline_result, tmp_path):
        path = tmp_path / "test_report.xlsx"
        write_excel_report(pipeline_result, path)
        wb = load_workbook(path)
        ws = wb["Report Info"]
        assert ws["A1"].value == "Transaction Analysis Report"

    def test_analysis_sheets_have_headers(self, pipeline_result, tmp_path):
        path = tmp_path / "test_report.xlsx"
        write_excel_report(pipeline_result, path)
        wb = load_workbook(path)
        # Check first analysis sheet (skip Report Info and Contents)
        if len(wb.sheetnames) > 2:
            ws = wb[wb.sheetnames[2]]
            assert ws.cell(row=1, column=1).value is not None

    def test_frozen_panes(self, pipeline_result, tmp_path):
        path = tmp_path / "test_report.xlsx"
        write_excel_report(pipeline_result, path)
        wb = load_workbook(path)
        if len(wb.sheetnames) > 2:
            ws = wb[wb.sheetnames[2]]
            assert ws.freeze_panes == "A2"

    def test_skips_failed_analyses(self, sample_csv_path, tmp_path):
        import pandas as pd

        from txn_analysis.analyses.base import AnalysisResult

        settings = Settings.from_args(data_file=sample_csv_path, output_dir=tmp_path)
        result = PipelineResult(
            settings=settings,
            df=pd.DataFrame(),
            analyses=[
                AnalysisResult.from_df("ok", "OK", pd.DataFrame({"a": [1]}), sheet_name="OK"),
                AnalysisResult.from_df("bad", "Bad", pd.DataFrame(), error="boom"),
            ],
        )
        path = tmp_path / "test_report.xlsx"
        write_excel_report(result, path)
        wb = load_workbook(path)
        sheet_names = wb.sheetnames
        assert "OK" in sheet_names
        assert "bad" not in sheet_names
