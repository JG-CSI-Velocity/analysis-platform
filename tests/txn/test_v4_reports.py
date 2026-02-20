"""Tests for txn_analysis.v4_html_report and v4_excel_report."""

from __future__ import annotations

import pandas as pd
import plotly.graph_objects as go
import pytest

from txn_analysis.v4_excel_report import (
    CURRENCY_FORMAT,
    HEADER_FILL,
    HEADER_FONT,
    NUMBER_FORMAT,
    PERCENT_FORMAT,
    format_df_for_excel,
    generate_excel_report,
)
from txn_analysis.v4_html_report import build_kpi_html, generate_html_report

# ---------------------------------------------------------------------------
# Shared fixtures
# ---------------------------------------------------------------------------


@pytest.fixture()
def sample_storyline_results():
    """Minimal storyline results dict for report generation."""
    fig = go.Figure(go.Bar(x=["A", "B"], y=[10, 20]))
    df = pd.DataFrame({"Metric": ["Spend", "Count"], "Value": [1000, 50]})
    return {
        "s1_portfolio": {
            "title": "S1: Portfolio Health",
            "description": "Monthly trends and activation",
            "sections": [
                {
                    "heading": "KPI Summary",
                    "narrative": "<b>Strong growth</b>",
                    "figures": [fig],
                    "tables": [("Monthly Summary", df)],
                }
            ],
            "sheets": [
                {
                    "name": "S1 Monthly",
                    "df": df,
                    "currency_cols": ["Value"],
                    "pct_cols": [],
                    "number_cols": [],
                }
            ],
        },
        "s6_risk": {
            "title": "S6: Risk",
            "description": "Balance tiers",
            "sections": [],
            "sheets": [],
        },
    }


@pytest.fixture()
def sample_config():
    return {"client_name": "Test CU", "client_id": "1234"}


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------


class TestExcelReportConstants:
    def test_header_fill_color(self):
        assert HEADER_FILL.start_color.rgb == "002E4057"

    def test_header_font_bold(self):
        assert HEADER_FONT.bold is True

    def test_currency_format(self):
        assert CURRENCY_FORMAT == "#,##0"

    def test_percent_format(self):
        assert PERCENT_FORMAT == '0.0"%"'

    def test_number_format(self):
        assert NUMBER_FORMAT == "#,##0"


class TestGenerateExcelReport:
    def test_creates_file(self, tmp_path, sample_storyline_results, sample_config):
        out = tmp_path / "report.xlsx"
        generate_excel_report(sample_storyline_results, sample_config, str(out))
        assert out.exists()
        assert out.stat().st_size > 0

    def test_has_overview_sheet(self, tmp_path, sample_storyline_results, sample_config):
        from openpyxl import load_workbook

        out = tmp_path / "report.xlsx"
        generate_excel_report(sample_storyline_results, sample_config, str(out))
        wb = load_workbook(out)
        assert "Overview" in wb.sheetnames
        assert "S1 Monthly" in wb.sheetnames

    def test_skips_empty_sheets(self, tmp_path, sample_config):
        results = {
            "s1": {
                "title": "S1",
                "sheets": [{"name": "Empty", "df": pd.DataFrame()}],
            }
        }
        out = tmp_path / "report.xlsx"
        generate_excel_report(results, sample_config, str(out))
        from openpyxl import load_workbook

        wb = load_workbook(out)
        assert "Empty" not in wb.sheetnames


class TestFormatDfForExcel:
    def test_converts_currency_strings(self):
        df = pd.DataFrame({"amount": ["$1,234", "$5,678"]})
        result = format_df_for_excel(df)
        assert pd.api.types.is_numeric_dtype(result["amount"])

    def test_converts_percent_strings(self):
        df = pd.DataFrame({"rate": ["45.2%", "67.8%"]})
        result = format_df_for_excel(df)
        assert pd.api.types.is_numeric_dtype(result["rate"])

    def test_leaves_plain_text_alone(self):
        df = pd.DataFrame({"name": ["Alice", "Bob"]})
        result = format_df_for_excel(df)
        assert result["name"].tolist() == ["Alice", "Bob"]

    def test_does_not_modify_original(self):
        df = pd.DataFrame({"amount": ["$100"]})
        result = format_df_for_excel(df)
        assert df["amount"].iloc[0] == "$100"


# ---------------------------------------------------------------------------
# HTML report
# ---------------------------------------------------------------------------


class TestBuildKpiHtml:
    def test_basic_kpi(self):
        kpis = [{"label": "Spend", "value": "$1.2M"}]
        html = build_kpi_html(kpis)
        assert "kpi-card" in html
        assert "$1.2M" in html
        assert "Spend" in html

    def test_kpi_with_positive_change(self):
        kpis = [{"label": "Growth", "value": "15%", "change": 5.0}]
        html = build_kpi_html(kpis)
        assert "positive" in html
        assert "5.0%" in html

    def test_kpi_with_negative_change(self):
        kpis = [{"label": "Churn", "value": "8%", "change": -3.2}]
        html = build_kpi_html(kpis)
        assert "negative" in html
        assert "3.2%" in html

    def test_multiple_kpis(self):
        kpis = [
            {"label": "A", "value": "1"},
            {"label": "B", "value": "2"},
            {"label": "C", "value": "3"},
        ]
        html = build_kpi_html(kpis)
        assert html.count("kpi-card") == 3


class TestGenerateHtmlReport:
    def test_creates_file(self, tmp_path, sample_storyline_results, sample_config):
        out = tmp_path / "dashboard.html"
        generate_html_report(sample_storyline_results, sample_config, str(out))
        assert out.exists()
        content = out.read_text()
        assert "Test CU" in content
        assert "S1: Portfolio Health" in content
        assert "plotly" in content.lower()

    def test_nav_links(self, tmp_path, sample_storyline_results, sample_config):
        out = tmp_path / "dashboard.html"
        generate_html_report(sample_storyline_results, sample_config, str(out))
        content = out.read_text()
        assert 'href="#s1_portfolio"' in content
        assert 'href="#s6_risk"' in content

    def test_sections_rendered(self, tmp_path, sample_storyline_results, sample_config):
        out = tmp_path / "dashboard.html"
        generate_html_report(sample_storyline_results, sample_config, str(out))
        content = out.read_text()
        assert "KPI Summary" in content
        assert "Strong growth" in content
